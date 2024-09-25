# -*- coding: utf-8 -*-
"""
Created on Tue Sep  3 08:33:16 2024

@author: Mayra Herrera
"""

import pandas as pd
import openpyxl 
import os
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

#Variable para porcentaje global
global porcentajeTotal 
porcentajeTotal = 100

#Importación del archivo excel para utilizar sus datos
try:
    url = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/pruebas_proc.xlsx'
    dfs = pd.read_excel(url, sheet_name=None, header=None)
except Exception as e:
    print(f'Error al cargar el archivo de {url} verifique la direccion: {str(e)}')
    
try:
    url = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
    dfs2 = pd.read_excel(url, sheet_name=None, header=None)
except Exception as e:
    print(f'Error al cargar el archivo de {url} verifique la direccion: {str(e)}')
    
general =dfs['Hoja1']
general.columns=general.iloc[0]
general = general[1:]
general.reset_index(drop=True, inplace=True)

cliente = dfs2['TablaGeneral']
cliente.columns=cliente.iloc[0]
cliente = cliente[1:]
cliente.reset_index(drop=True, inplace=True)

procedimiento = input('Ingrese el procedimiento que desea desglozar: ')
clave = set(general[general['N° Procedimiento']==procedimiento]['Clave'])
clave = [cl for cl in clave]

filas = cliente[(cliente['Procedimiento/Licitación']==procedimiento)&(cliente['Clave']==clave[0])]
filas2 = general[(general['N° Procedimiento']==procedimiento)&(general['Clave']==clave[0])]

cantidad = filas['Cantidad']
cantidad = int(cantidad.iloc[0])

mes = filas['Mes']
mes = str(mes.iloc[0])

instituciones = set(filas2['Institución GRAL'])
instituciones = [i for i in instituciones]
cantidadInstituciones = len(instituciones)

subInstituciones = []
diccInstituciones = {}

for i in range(len(instituciones)):
    registro = general[(general['N° Procedimiento']==procedimiento)&(general['Clave']==clave[0]) & (general['Institución GRAL']==instituciones[i])]
    sub = registro['Institución Detalle']
    sub = [i for i in sub]
    subInstituciones.append(sub)
    diccInstituciones[instituciones[i]] = sub 

totalInstituciones = filas2['Institución Detalle']
totalInstituciones = [i for i in totalInstituciones]
cantidadTotalInstituciones = len(totalInstituciones)

diccPorcentajesUno = {}
for i in diccInstituciones.keys():
    diccPorcentajesUno[i] = (porcentajeTotal/len(diccInstituciones))/100
    

diccPorcentajesDos = {}
for i in diccInstituciones.keys():
    for j in diccInstituciones[i]:
        diccPorcentajesDos[j]=(diccPorcentajesUno[i]/len(diccInstituciones[i]))/100
    
diccCantidades = {}
for i in diccPorcentajesDos.keys():
    diccCantidades[i]=round((cantidad*diccPorcentajesDos[i])*100)
    
diccPorcentajeGeneral = {}
for i in diccPorcentajesDos.keys():
    diccPorcentajeGeneral[i] = (diccCantidades[i]/cantidad)
    
# Cargar el archivo de Excel existente
file_path = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/pruebas_proc.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb['Hoja1']

new_data = {
    'Porcentaje Por Institución': [],
    'Porcentaje General': [],
    'Cantidad': [],
    'Mes':[]
}

for index, row in general.iterrows():
    institucion_detalle = row['Institución Detalle']
    
    # Buscar el porcentaje y cantidad correspondiente
    if institucion_detalle in diccPorcentajesDos:
        new_data['PorcentajeUno'].append(diccPorcentajesUno[row['Institución GRAL']])
        new_data['PorcentajeDos'].append(diccPorcentajesDos[institucion_detalle])
        new_data['Cantidad'].append(diccCantidades[institucion_detalle])
    else:
        # Si el detalle no está en los diccionarios, agregar valores nulos
        new_data['PorcentajeUno'].append(None)
        new_data['PorcentajeDos'].append(None)
        new_data['Cantidad'].append(None)

# Agregar nuevas columnas al DataFrame
for col_name, col_data in new_data.items():
    general[col_name] = col_data

# Guardar el DataFrame actualizado en el archivo Excel
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    general.to_excel(writer, sheet_name='Hoja1', index=False)

# Guardar los cambios en el archivo de Excel
wb.save(file_path)





