# -*- coding: utf-8 -*-
"""
Created on Fri Aug 23 08:29:56 2024

@author: Mayra Herrera
"""

import tkinter as tk
import pandas as pd
from collections import Counter
from tkinter import ttk
from datetime import datetime
import openpyxl 
import os
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

######################################################################################
############################PRIMERA VENTANA QUE VE EL USUARIO#########################

class VentanaPrincipal(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Ventana de Usuario")
        self.geometry("600x400")
        
######################################################################################
#########################ESPACIOS PARA CARGAR LA BASE DE DATOS########################

        #Importación del archivo excel para utilizar sus datos
        try:
            url = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/PROCEDIMIENTOS.xlsx'
            self.dfs = pd.read_excel(url, sheet_name=None, header=None)
        except Exception as e:
            print(f'Error al cargar el archivo de {url} verifique la direccion: {str(e)}')
            
        try:
            url2 = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Laboratorios Vanquish/1.- BDL&MIBI/3.- MIBI/Catálogos/20231005_Catalogo Productos_V0.1.xlsx'
            self.dfs2 = pd.read_excel(url2, sheet_name=None, header=None)
        except Exception as e:
            print(f'Error al cargar el archivo de {url} verifique la direccion: {str(e)}')
            
        #Arreglar el excel para su uso   
        self.df= self.dfs
        self.hojaProc = self.df['PROCEDIMIENTOS']
        self.hojaProc.columns=self.hojaProc.iloc[1]
        self.hojaProc = self.hojaProc[2:]
        self.hojaProc.reset_index(drop=True, inplace=True)
        self.hojaProc = self.hojaProc
        
        #Obtener la hoja de los distribuidores
        self.df2= self.dfs2
        self.hojaDist = self.df2['BD']
        self.hojaDist.columns=self.hojaDist.iloc[2]
        self.hojaDist = self.hojaDist[3:]
        self.hojaDist.reset_index(drop=True, inplace=True)
        self.hojaDist.drop(columns=self.hojaDist.columns[0],inplace=True)

        
######################################################################################
################################ESPACIOS PARA LOS ELEMENTOS###########################
        
        #Label de las instrucciones
        self.labelTitulo = tk.Label(self, text="Eliga la acción que desea realizar y \n el tipo de usuario que es")
        self.labelTitulo.place(x=200,y=30)
        
        #Seccion de los BUTTONS para elegir la acción que quiere el usuario
        
        #Boton para ingreso nuevo
        self.botonIngreso = tk.Button(self,text="Ingreso de nueva información",width=30,
                                      height=10, command=self.ingresarNuevosDatos)
        self.botonIngreso.place(x=50,y=110)
        
        #Boton para modificacion de información
        self.botonModificacion = tk.Button(self,text="Modificación de información",width=30,
                                      height=10,command=self.abrirVentanaModificacion)
        self.botonModificacion.place(x=320,y=110)
        
        #Sección de los RADIOBUTTONS para elegir el tipo de cliente
        
        #Variable para guardar el tipo de cliente 
        self.tipoCliente = tk.StringVar(self,'Usuario')
        
        #Primera opción de tipo de cliente PUBLICO
        self.usuarioPublico = tk.Radiobutton(self,text="Público",variable=self.tipoCliente,
                                             value="publico")
        self.usuarioPublico.place(x=90,y=320)
        
        #Segunda opción de tipo de cliente DISTRIBUIDOR
        self.usuarioDistribuidor = tk.Radiobutton(self,text="Distribuidor",variable=self.tipoCliente,
                                             value="distribuidor")
        self.usuarioDistribuidor.place(x=260,y=320)
        
        #Tercera opción de tipo de cliente RETAIL
        self.usuarioRetail = tk.Radiobutton(self,text="Retail",variable=self.tipoCliente,
                                             value="retail")
        self.usuarioRetail.place(x=430,y=320)
        
#########################################################################################
#################################ESPACIO PARA LAS FUNCIONES##############################
    
    #Función para obtener el tipo de cliente que eligio el usuario
    def obtenerTipoDeCliente(self):
        if self.tipoCliente.get() == 'publico':
            cliente = 'publico'
        elif self.tipoCliente.get() == 'distribuidor':
            cliente = 'distribuidor'
        else:
            cliente = 'retail'
            
        return cliente
    

    #Funcion para ir a la ventana de ingreso de información 
    def ingresarNuevosDatos(self):
        if self.verificarEleccionUsuario():
            return
        
        self.cliente = self.obtenerTipoDeCliente()
        if self.cliente == 'publico':
            self.abrirVentanaPublicoNuevo()
        else:
            self.abrirVentanaDistribuidoroRetailNuevo()
    
    def verificarEleccionUsuario(self):
        if self.tipoCliente.get() == 'Usuario':
            tk.messagebox.showwarning(title=None, message='Eliga un tipo de usuario')
            return True
        return False
            
    #Función para abrir la ventana de registro de cliente publico
    def abrirVentanaPublicoNuevo(self):
        print(self.cliente)
        self.ventana2 = VentanaPublico(self)
        self.withdraw()
    
    #Función para abrir la ventana de registro de cliente distribuidor o retail
    def abrirVentanaDistribuidoroRetailNuevo(self):
        print(self.cliente)
        self.ventana2 = VentanaDistRet(self)
        self.withdraw()
    
    #Función para abrir la ventana de modificacion de informacion
    def abrirVentanaModificacion(self):
        if self.verificarEleccionUsuario():
            return
        
        cliente = self.obtenerTipoDeCliente()
        print(cliente)
        self.ventana2 = VentanaModificacion(self)
        self.withdraw()

######################################################################################
############################VENTANA DE INGRESO PARA PUBLICO###########################        

class VentanaPublico(tk.Toplevel):
    def __init__(self,VentanaPrincipal):
        super().__init__()
        self.title("Ventana Ingreso Publico")
        self.geometry("500x650")
        self.Ventana_Principal = VentanaPrincipal
        
######################################################################################
#########################ESPACIOS PARA CARGAR LA BASE DE DATOS########################

        self.hojaProcedimientos = self.Ventana_Principal.hojaProc
        tamaño = self.hojaProcedimientos.shape
        print(tamaño)
        

######################################################################################
################################ESPACIOS PARA LOS ELEMENTOS###########################

        #Label para la ventana de tipo cliente publico
        self.labelMensajePublico = tk.Label(self,text='Ingreso de datos del tipo de cliente PÚBLICO')
        self.labelMensajePublico.place(x=130,y=10)
        
        #Seccion de los ENTRY y LISTBOX para ingresar la información del cliente
        
        #Variable para el parametro de procedimiento
        self.procedimientoPublico = tk.StringVar(self,'')     
        #Texto para el procedimiento
        self.labelProc = tk.Label(self,text='Procedimiento/Licitacion')
        self.labelProc.place(x=20,y=130)
        #Entrada de texto para el procedimiento
        self.entradaProc = tk.Entry(self, width =40, textvariable=self.procedimientoPublico,
                                          state=tk.DISABLED)
        self.entradaProc.place(x=175,y=130)
        
        #Variable para el parametro de Clave
        self.clavePublico = tk.StringVar(self,'')     
        #Texto para la clave
        self.labelClave = tk.Label(self,text='Clave')
        self.labelClave.place(x=20,y=170)
        #Entrada de texto para la clave
        self.entradaClave = tk.Entry(self, width =40, textvariable=self.clavePublico,
                                          state=tk.DISABLED)
        self.entradaClave.place(x=175,y=170)
        
        
        #Cuadro de texto para mostrar las claves o procedimientos encontrados
        self.scroll1 = tk.Scrollbar(self, orient=tk.VERTICAL)
        self.listaDatosPublico = tk.Listbox(self, width=30, height=4, yscrollcommand=self.scroll1.set,
                                            selectmode=tk.SINGLE,exportselection=False)
        self.scroll1.configure(command=self.listaDatosPublico.yview)                 
        self.listaDatosPublico.place(x=175, y=210)
        self.scroll1.place(x=190, y=210)
        
        #Variable para el parametro de cantidad
        self.cantidad=tk.StringVar(self,'0')
        #Texto para la cantidad
        self.labelCantidad = tk.Label(self,text='Cantidad')
        self.labelCantidad.place(x=20,y=290)
        #Entrada de texto para la cantidad
        self.entradaCantidad = tk.Entry(self, width =40, textvariable=self.cantidad,
                                          state=tk.NORMAL)
        self.entradaCantidad.place(x=175,y=290)
        
        #Texto para el mes
        self.labelMes = tk.Label(self,text='Mes/Meses')
        self.labelMes.place(x=20,y=330)
        
        ########################################################################################################
        #Cuadro de texto para mostrar los meses del año
        self.scroll2 = tk.Scrollbar(self, orient=tk.VERTICAL)
        self.listaMesesPublico = tk.Listbox(self, width=30, height=4,selectmode=tk.MULTIPLE,
                                            yscrollcommand=self.scroll2.set,exportselection=False)
        self.scroll2.configure(command=self.listaMesesPublico.yview) 
        self.listaMesesPublico.place(x=175, y=330)
        self.scroll2.place(x=190, y=330)
        ########################################################################################################
        
        
        # Cuadro de texto para mostrar el resumen de los datos ingresados
        self.texto = tk.Text(self, height=5, width=40, wrap='word')
        self.texto.insert(tk.END, "")
        self.texto.config(state='disabled')  # Configurar el widget como solo lectura
        self.texto.place(x=90,y=440)
        
        
        #Seccion de los RADIOBUTTONS para elegir la clave o procedimiento
        
        #Radiobutton para que el usuario publico eliga si licitacion o clave para ingresar
        #Variable para la opción elegida de busqueda por el usuario
        self.datoIngreso = tk.StringVar(self,'ninguno')
        #Eleccion Licitacion
        self.eleccionProc = tk.Radiobutton(self,text="",variable=self.datoIngreso,
                                             value="procedimiento",command=self.habilitacionDeEntradas)
        self.eleccionProc.place(x=450,y=130)
        #Eleccion clave
        self.eleccionClave = tk.Radiobutton(self,text="",variable=self.datoIngreso,
                                             value="clave",command=self.habilitacionDeEntradas)
        self.eleccionClave.place(x=450,y=170)
        
        #Radiobutton para que el usuario publico eliga si quiere varios meses o todo el año
        #Variable para la opción elegida de meses por el usuario
        self.datoMes = tk.StringVar(self,'ninguno')
        #Eleccion Licitacion
        self.eleccionMesUnico = tk.Radiobutton(self,text="Por meses",variable=self.datoMes,
                                             value="unicoMes",command=self.habilitacionMeses)
        self.eleccionMesUnico.place(x=375,y=330)
        #Eleccion clave
        self.eleccionCompleto = tk.Radiobutton(self,text="Año \ncompleto",variable=self.datoMes,
                                             value="añoCompleto",command=self.habilitacionMeses)
        self.eleccionCompleto.place(x=375,y=360)
        
        
        #Seccion de los BUTTONS para realizar acciones dentro de la ventana
        
        #Boton para mostrar las claves o procedimientos encontrados
        self.mostrarClaves = tk.Button(self, text="Mostrar Claves/Procedimientos",
                                       width=15,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.mostrarClaProc)
        self.mostrarClaves.place(x=370, y=210)
        
        #Boton para mostrar el resumen de datos ingresados
        self.mostrarResumen = tk.Button(self, text="Mostrar resumen de ingreso",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.visualizarResumenEntrada)
        self.mostrarResumen.place(x=40, y=550)
        
        #Boton para guardar la información en el excel de registros
        self.guardarRegistro = tk.Button(self, text="Guardar información",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.guardarInformacion)
        self.guardarRegistro.place(x=185, y=550)
        
        #Boton para regresar a la página principal
        self.regresoPublicoInicio = tk.Button(self, text="Regresar al Inicio",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.volverVentanaPrincipal)
        self.regresoPublicoInicio.place(x=330, y=550)
        
        #Seccion de los CHECKBUTTONS para elegir los meses
        #Mes de ENERO
        # self.mesEnero = tk.StringVar(self,'')
        # self.checkEnero = tk.Checkbutton(self,text='ENERO',variable=self.mesEnero,onvalue='ENERO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkEnero.place(x=100,y=330)
        
        # #Mes de FEBRERO
        # self.mesFebrero = tk.StringVar(self,'')
        # self.checkFebrero = tk.Checkbutton(self,text='FEBRERO',variable=self.mesFebrero,onvalue='FEBRERO',offvalue='',
        #                                    state=tk.DISABLED)
        # self.checkFebrero.place(x=180,y=330)
        
        # #Mes de MARZO
        # self.mesMarzo = tk.StringVar(self,'')
        # self.checkMarzo = tk.Checkbutton(self,text='MARZO',variable=self.mesMarzo,onvalue='MARZO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkMarzo.place(x=270,y=330)
        
        # #Mes de ABRIL
        # self.mesAbril = tk.StringVar(self,'')
        # self.checkAbril = tk.Checkbutton(self,text='ABRIL',variable=self.mesAbril,onvalue='ABRIL',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkAbril.place(x=100,y=350)
        
        # #Mes de MAYO
        # self.mesMayo = tk.StringVar(self,'')
        # self.checkMayo = tk.Checkbutton(self,text='MAYO',variable=self.mesMayo,onvalue='MAYO',offvalue='',
        #                                 state=tk.DISABLED)
        # self.checkMayo.place(x=180,y=350)
        
        # #Mes de JUNIO
        # self.mesJunio = tk.StringVar(self,'')
        # self.checkJunio = tk.Checkbutton(self,text='JUNIO',variable=self.mesJunio,onvalue='JUNIO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkJunio.place(x=270,y=350)
        
        # #Mes de JULIO
        # self.mesJulio = tk.StringVar(self,'')
        # self.checkJulio = tk.Checkbutton(self,text='JULIO',variable=self.mesJulio,onvalue='JULIO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkJulio.place(x=100,y=370)
        
        # #Mes de AGOSTO
        # self.mesAgosto = tk.StringVar(self,'')
        # self.checkAgosto = tk.Checkbutton(self,text='AGOSTO',variable=self.mesAgosto,onvalue='AGOSTO',offvalue='',
        #                                   state=tk.DISABLED)
        # self.checkAgosto.place(x=180,y=370)
        
        # #Mes de SEPTIEMBRE
        # self.mesSeptiembre = tk.StringVar(self,'')
        # self.checkSeptiembre = tk.Checkbutton(self,text='SEPTIEMBRE',variable=self.mesSeptiembre,onvalue='SEPTIEMBRE',offvalue='',
        #                                       state=tk.DISABLED)
        # self.checkSeptiembre.place(x=270,y=370)
        
        # #Mes de OCTUBRE
        # self.mesOctubre = tk.StringVar(self,'')
        # self.checkOctubre = tk.Checkbutton(self,text='OCTUBRE',variable=self.mesOctubre,onvalue='OCTUBRE',offvalue='',
        #                                    state=tk.DISABLED)
        # self.checkOctubre.place(x=100,y=390)
        
        # #Mes de NOVIEMBRE
        # self.mesNoviembre = tk.StringVar(self,'')
        # self.checkNoviembre = tk.Checkbutton(self,text='NOVIEMBRE',variable=self.mesNoviembre,onvalue='NOVIEMBRE',offvalue='',
        #                                      state=tk.DISABLED)
        # self.checkNoviembre.place(x=180,y=390)
        
        # #Mes de DICIEMBRE
        # self.mesDiciembre = tk.StringVar(self,'')
        # self.checkDiciembre = tk.Checkbutton(self,text='DICIEMBRE',variable=self.mesDiciembre,onvalue='DICIEMBRE',offvalue='',
        #                                      state=tk.DISABLED)
        # self.checkDiciembre.place(x=270,y=390)
        
        
#########################################################################################
#################################ESPACIO PARA LAS FUNCIONES##############################

    def mostrarClaProc(self):       
        self.parametroBusqueda = self.capturaDeParametroBusqueda()
        print(self.parametroBusqueda)
        
        if self.datoIngreso.get() == 'procedimiento':
            if not self.parametroBusqueda:
                tk.messagebox.showwarning(title=None, message='Por favor ingrese una licitación.')
                return
            clavesEncontradas = self.obtenerClavesProcedimientos(self.parametroBusqueda)
            self.listaDatosPublico.delete(0, tk.END)
            
            # Agregar las claves a la lista
            for clave in clavesEncontradas:
                self.listaDatosPublico.insert(tk.END, clave)
        else:
            if not self.parametroBusqueda:
                tk.messagebox.showwarning(title=None, message='Por favor ingrese una clave')
                return
        
            procedimientosEncontrados = self.obtenerClavesProcedimientos(self.parametroBusqueda)
            
            self.listaDatosPublico.delete(0, tk.END)
            
            # Agregar las claves a la lista
            for procedimiento in procedimientosEncontrados:
                self.listaDatosPublico.insert(tk.END, procedimiento)
        
    def obtenerClavesProcedimientos(self, pBusqueda):
        
        if self.datoIngreso.get() == 'procedimiento':
            self.filasProc = self.hojaProcedimientos[self.hojaProcedimientos['N° Procedimiento'] == pBusqueda]
            clavesProc = self.filasProc['Clave'].tolist()
            clavesProc = set(clavesProc)
            return clavesProc
        else:
            self.filasProc = self.hojaProcedimientos[self.hojaProcedimientos['Clave'] == pBusqueda]
            ProcClav = self.filasProc['N° Procedimiento'].tolist()
            ProcClav = set(ProcClav)
            return ProcClav
        
    def obtenerPrecio(self, dElegido):
        print(self.parametroBusqueda)
        
        if self.datoIngreso.get() == 'procedimiento':
            self.filasClav = self.hojaProcedimientos[(self.hojaProcedimientos['Clave'] == dElegido) & (self.hojaProcedimientos['N° Procedimiento']==self.parametroBusqueda)]
            precioClav = self.filasClav['P.U.'].tolist()
            print(precioClav)
            contador = Counter(precioClav)
            precioClav,frecuencia = contador.most_common(1)[0]
        else:
            self.filasClav = self.hojaProcedimientos[(self.hojaProcedimientos['Clave'] == self.parametroBusqueda) & (self.hojaProcedimientos['N° Procedimiento']==dElegido)]
            precioClav = self.filasClav['P.U.'].tolist()
            print(precioClav)
            contador = Counter(precioClav)
            precioClav,frecuencia = contador.most_common(1)[0]

        return precioClav
      
    
    def capturaDeParametroBusqueda(self):
        if self.verificarParametroBusqueda():
            return
        
        if self.datoIngreso.get() == 'procedimiento':
            self.proc_str = self.procedimientoPublico.get()
            return self.proc_str 
        else:
            self.clave_str = self.clavePublico.get()
            return self.clave_str 
              
            
    def habilitacionDeEntradas(self):
        if self.datoIngreso.get() == 'procedimiento':
            self.entradaClave.config(state=tk.DISABLED)
            self.entradaProc.config(state=tk.NORMAL)
        else:
            self.entradaClave.config(state=tk.NORMAL)
            self.entradaProc.config(state=tk.DISABLED)
            
    def habilitacionMeses(self):
        ######################################################################################
        meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
        
        if self.datoMes.get() == 'unicoMes':
            self.listaMesesPublico.config(state=tk.NORMAL)
            self.listaMesesPublico.delete(0, tk.END)
                     
            # Agregar las claves a la lista
            for mes in meses:
                self.listaMesesPublico.insert(tk.END, mes)
                
        else:
            self.listaMesesPublico.config(state=tk.DISABLED)
        #######################################################################################
        
        # if self.datoMes.get() == 'unicoMes':
        #     self.checkEnero.config(state=tk.NORMAL)
        #     self.checkFebrero.config(state=tk.NORMAL)
        #     self.checkMarzo.config(state=tk.NORMAL)
        #     self.checkAbril.config(state=tk.NORMAL)
        #     self.checkMayo.config(state=tk.NORMAL)
        #     self.checkJunio.config(state=tk.NORMAL)
        #     self.checkJulio.config(state=tk.NORMAL)
        #     self.checkAgosto.config(state=tk.NORMAL)
        #     self.checkSeptiembre.config(state=tk.NORMAL)
        #     self.checkOctubre.config(state=tk.NORMAL)
        #     self.checkNoviembre.config(state=tk.NORMAL)
        #     self.checkDiciembre.config(state=tk.NORMAL)
        # else:
        #     self.checkEnero.config(state=tk.DISABLED)
        #     self.checkFebrero.config(state=tk.DISABLED)
        #     self.checkMarzo.config(state=tk.DISABLED)
        #     self.checkAbril.config(state=tk.DISABLED)
        #     self.checkMayo.config(state=tk.DISABLED)
        #     self.checkJunio.config(state=tk.DISABLED)
        #     self.checkJulio.config(state=tk.DISABLED)
        #     self.checkAgosto.config(state=tk.DISABLED)
        #     self.checkSeptiembre.config(state=tk.DISABLED)
        #     self.checkOctubre.config(state=tk.DISABLED)
        #     self.checkNoviembre.config(state=tk.DISABLED)
        #     self.checkDiciembre.config(state=tk.DISABLED)
            
    def verificarParametroBusqueda(self):
        if self.datoIngreso.get() == '':
            tk.messagebox.showwarning(title=None, message='Elija un parametro de busqueda')
            return True
        return False
    
    def verificarParametroTiempo(self):
        if self.datoMes.get() == 'ninguno':
            tk.messagebox.showwarning(title=None, message='Elija un tipo de tiempo')
            return True
        return False
    
    def verificarIntegridadDatos(self):
        
        if self.datoIngreso.get()=='ninguno' or self.datoMes.get() == 'ninguno':
            tk.messagebox.showwarning(title=None, message='Llene toda la información ')
            return
        elif self.datoIngreso.get()=='procedimiento' and self.procedimientoPublico.get()=='':
            tk.messagebox.showwarning(title=None, message='Ingrese un procedimiento o licitación')
            return
        elif self.datoIngreso.get()=='clave' and self.clavePublico.get()=='':
            tk.messagebox.showwarning(title=None, message='Ingrese una clave')
            return
        
        # Obtener la clave o procedimiento seleccionada
        seleccion = self.listaDatosPublico.curselection()
        #Verificar que si se ha seleccionado una clave
        if not seleccion:
            tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
            return
                  
        if self.cantidad.get()!='0' or self.cantidad.get()=='0':
            try:
                self.cantidadEntero = int(self.cantidad.get())
            except ValueError:
                tk.messagebox.showwarning(title=None, message='El tipo de dato en Cantidad es erróneo')
                return
        else:       
            tk.messagebox.showwarning(title=None, message='Llene toda la información ')
            return
        ######################################################################################################
        if self.datoMes.get()=='unicoMes':
            # Obtener los meses seleccionados
            seleccionMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
            #Verificar que si se ha seleccionado una clave
            if not seleccionMes:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                return
        #######################################################################################################
        
        # if self.datoMes.get()=='unicoMes':
        #     #verificar si hay meses seleeccionados
        #     checks = [self.mesEnero.get(),self.mesFebrero.get(),self.mesMarzo.get(), self.mesAbril.get(),
        #               self.mesMayo.get(),self.mesJunio.get(),self.mesJulio.get(),self.mesAgosto.get(),
        #               self.mesSeptiembre.get(),self.mesOctubre.get(),self.mesNoviembre.get(),
        #               self.mesDiciembre.get()]
            
        #     todosChecks = all(val=='' for val in checks)
            
        #     if todosChecks:
        #         tk.messagebox.showwarning(title=None, message='Seleccione al menos un mes')
        #         return
        
        return True
    
    def obtencionDatosResumen(self):
        
        if self.verificarIntegridadDatos():
        
            # Obtener la clave o procedimiento seleccionada
            seleccion = self.listaDatosPublico.curselection()
            #Verificar que si se ha seleccionado una clave
            if not seleccion:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                return 
            
            #Obtener el procedimiento y clave
            if self.datoIngreso.get()=='procedimiento':
                self.datoProcedimiento = self.procedimientoPublico.get()
                #Obtener la clave seleccionada 
                self.datoClave = self.listaDatosPublico.get(seleccion[0])
    
            else:
                #Obtener la clave seleccionada 
                self.datoProcedimiento = self.listaDatosPublico.get(seleccion[0])
                self.datoClave = self.clavePublico.get()
            
            #Obtener el precio publico
            if self.datoIngreso.get() == 'procedimiento':
                self.datoPrecio = self.obtenerPrecio(self.datoClave)
            else:
                self.datoPrecio = self.obtenerPrecio(self.datoProcedimiento)
            
            #Obtener la cantidad
            self.datoCantidad = self.cantidad.get()
            
            if self.verificarParametroTiempo():
                return
            #########################################################################################################
            #Obtener mes o meses 
            if self.datoMes.get()=='unicoMes':
                #Obtener la clave seleccionada 
                self.varMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
                if not self.varMes:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                    return
            else:
                self.varMes = self.datoMes.get()
            #########################################################################################################
            # if self.datoMes.get()=='unicoMes':
            #     #Obtener los meses seleccionados
            #     checks = [self.mesEnero.get(),self.mesFebrero.get(),self.mesMarzo.get(), self.mesAbril.get(),
            #               self.mesMayo.get(),self.mesJunio.get(),self.mesJulio.get(),self.mesAgosto.get(),
            #               self.mesSeptiembre.get(),self.mesOctubre.get(),self.mesNoviembre.get(),
            #               self.mesDiciembre.get()]
                
            #     self.varMes = [val for val in checks if val != '']
                
            # else:
            #     self.varMes = self.datoMes.get()
                
                       
            return self.datoProcedimiento,self.datoClave,self.datoPrecio,self.datoCantidad,self.varMes

            
    def visualizarResumenEntrada(self):
        proc,cla,pre,can,me=self.obtencionDatosResumen()
        
        self.texto.config(state=tk.NORMAL)
        self.texto.delete(1.0, tk.END)
        self.texto.insert(tk.END, f'Licitación/Procedimiento: {proc} \n')
        self.texto.insert(tk.END, f'Clave: {cla} \n')        
        self.texto.insert(tk.END, f'Cantidad: {can} \n')
        self.texto.insert(tk.END, f'Precio: ${pre} \n')
        if self.datoMes.get()=='unicoMes':
            textoMes = ', '.join(me)
        else:
            textoMes = me
        self.texto.insert(tk.END, f'Mes/Meses: {textoMes} \n')
        self.texto.config(state=tk.DISABLED)
        
    def crearTablaParaGuardar(self):
        proc,cla,pre,can,me=self.obtencionDatosResumen()
        fecha = datetime.now()
        year = fecha.year
        if me =='añoCompleto':
            meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                    'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
            dicc = { 'Mes':meses, 'Año':year,'Clave':cla,'Procedimiento/Licitación':proc,
                    'Cantidad':can,'Precio':pre,'Sector':'Público'}
        else:
            dicc = { 'Mes':me, 'Año':year,'Clave':cla,'Procedimiento/Licitación':proc,
                    'Cantidad':can,'Precio':pre,'Sector':'Público'}
        
        dataFrame = pd.DataFrame(dicc)
        print(dataFrame)
        return dataFrame
    
    def guardarInformacionExcel(self):
        dataFrame = self.crearTablaParaGuardar()
        
        try:
            ruta = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
            archivo = os.path.isfile(ruta)
            if archivo:
                with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    libro = pd.ExcelFile(ruta, engine='openpyxl')
                    if 'TablaGeneral' in libro.sheet_names:
                        dfExistente = pd.read_excel(ruta, sheet_name='TablaGeneral')
                        dfCombinado = pd.concat([dfExistente, dataFrame], ignore_index=True)
                        
                        # Eliminar duplicados
                        dfDuplicado = dfCombinado.drop_duplicates(subset=['Mes', 'Clave', 'Procedimiento/Licitación'], keep='last')
                        
                        #Filtrar por sector 
                        sector='Público'
                        dfSectorFiltrado = dfCombinado[dfCombinado['Sector']==sector]
                        
                        #Verificar duplicados
                        duplicados = dfSectorFiltrado[dfSectorFiltrado.duplicated(subset=['Mes', 'Clave', 'Procedimiento/Licitación'], keep=False)]
                        if not duplicados.empty:
                            mensajesDuplicados = duplicados[['Mes', 'Clave', 'Procedimiento/Licitación']].drop_duplicates().to_string(index=False)
                            message = f'Los siguientes registros ya están en el excel:\n{mensajesDuplicados}\n y no se ha guardado la información'
                            tk.messagebox.showwarning(title=None, message=message)
                        else:
                            dfCombinado.to_excel(writer, sheet_name='TablaGeneral', index=False)
                            tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
                    else:
                        dataFrame.to_excel(writer, sheet_name='TablaGeneral', index=False)
                        writer.sheets['TablaGeneral'].header_row = 1
                        workbook = writer.book
                        worksheet = writer.sheets['TablaGeneral']
                        
                        # Aplicar formato a los encabezados
                        for cell in worksheet[1]:
                            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
            else:
                dataFrame.to_excel(ruta, sheet_name='TablaGeneral', index=False)
                wb = load_workbook(ruta)
                ws = wb.active
                
                # Definir estilos para los encabezados
                header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                # Aplicar formato a los encabezados
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(ruta)
                tk.messagebox.showinfo(title=None, message='Tabla Creada e Información guardada con éxito')
        except ValueError:
            tk.messagebox.showwarning(title=None, message='Hubo un error al guardar la información')
        
        # try:
        #     ruta = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
        #     archivo = os.path.isfile(ruta)
        #     if archivo:
        #         with pd.ExcelWriter(ruta,engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
        #             libro = pd.ExcelFile(ruta,engine='openpyxl')
        #             if 'TablaGeneral' in libro.sheet_names:
        #                 dfExistente = pd.read_excel(ruta,sheet_name='TablaGeneral')                        
        #                 dfCombinado = pd.concat([dfExistente,dataFrame], ignore_index=True)
        #                 print(dfCombinado)
        #                 sector='Público'
        #                 dfSectorFiltrado = dfCombinado[dfCombinado['Sector']==sector]
        #                 dfValoresDuplicados = dfSectorFiltrado[dfSectorFiltrado.duplicated(subset=['Mes','Clave','Procedimiento/Licitación'],keep=False)]
        #                 duplicadosDatos = [dfValoresDuplicados['Mes'].unique(),dfValoresDuplicados['Clave'].unique(),dfValoresDuplicados['Procedimiento/Licitación'].unique()]
        #                 mensajesDuplicados = []
        #                 for i in range(1):
        #                     for j in range(len(duplicadosDatos[i])):
        #                         mensajesDuplicados.append(duplicadosDatos[i][j]+" "+duplicadosDatos[i+1][0]+' '+duplicadosDatos[i+2][0])
                                
        #                 dfDuplicado = dfSectorFiltrado.drop_duplicates(subset=['Mes','Clave','Procedimiento/Licitación'], keep='first')
        #                 print(dfDuplicado)
        #                 if len(dfSectorFiltrado) == len(dfDuplicado):
        #                     dfDuplicado.to_excel(writer,sheet_name='TablaGeneral',index=False)
        #                     tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
        #                 else:
        #                     message = 'Los siguientes registros ya estan en el excel: \n'+'\n'.join(mensajesDuplicados)+'\n'+'y no se ha guardado la información'
        #                     tk.messagebox.showwarning(title=None, message=message)
        #             else:
        #                 dataFrame.to_excel(writer,sheet_name='TablaGeneral',index=False)
        #                 writer.sheets['TablaGeneral'].header_row = 1
        #                 workbook  = writer.book
        #                 worksheet = writer.sheets['TablaGeneral']
        #                 # Aplicar formato a los encabezados
        #                 for cell in worksheet[1]:
        #                     cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        #                     cell.font = Font(color="FFFFFF", bold=True)
        #                 tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
        #     else:
        #         dataFrame.to_excel(ruta,sheet_name='TablaGeneral',index=False)
        #         path = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
        #         wb = load_workbook(path)
        #         ws = wb.active
        #         # Definir estilos para los encabezados
        #         header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')  # Azul
        #         header_font = Font(color='FFFFFF', bold=True)  # Blanco y negrita
        #         # Aplicar formato a los encabezados
        #         for cell in ws[1]:
        #             cell.fill = header_fill
        #             cell.font = header_font
        #         wb.save(path)               
        #         tk.messagebox.showinfo(title=None, message='Tabla Creada e Información guardada con éxito')
        # except ValueError:
        #     tk.messagebox.showwarning(title=None, message='Hubo un error al guardar la información')
        #     return
            
            
    def guardarInformacion(self):
        self.verificarIntegridadDatos()
        self.guardarInformacionExcel()


    
    def volverVentanaPrincipal(self):
        self.destroy()
        self.Ventana_Principal.deiconify()
    

########################################################################################
###################VENTANA DE INGRESO PARA DISTRIBUIDOR O RETAIL########################      

class VentanaDistRet(tk.Toplevel):
    def __init__(self,VentanaPrincipal):
        super().__init__()
        self.title("Ventana Ingreso Distribuidor/Retail")
        self.geometry("500x700")
        self.Ventana_Principal = VentanaPrincipal
        
######################################################################################
#########################ESPACIOS PARA CARGAR LA BASE DE DATOS########################

        self.hojaDistribuidor = self.Ventana_Principal.hojaDist
        tamaño = self.hojaDistribuidor.shape
        print(tamaño)
        

######################################################################################
################################ESPACIOS PARA LOS ELEMENTOS###########################

        self.tipoCliente = self.Ventana_Principal.tipoCliente.get()
        if self.tipoCliente =='distribuidor':
            mensaje = 'Ingreso de datos del tipo de cliente DISTRIBUIDOR'
        else:
            mensaje = 'Ingreso de datos del tipo de cliente RETAIL'

        #Label para la ventana de tipo cliente distribuidor o Retail
        self.labelMensajePublico = tk.Label(self,text=mensaje)
        self.labelMensajePublico.place(x=130,y=10)
        
        #Seccion de los ENTRY, LISTBOX y OPTIONMENU para ingresar la información del cliente
        
        #Label de instrucción del primer paso
        self.labelMensajePublico = tk.Label(self,text='Paso 1: Elija una molécula')
        self.labelMensajePublico.place(x=50,y=50)
        
        #Variable para el parametro de molecula
        self.moleculaDist = tk.StringVar(self,'Elija una molécula') 
        self.moleculaDist.trace_add('write', self.mostrarPresentaciones)
        #Texto para la molecula
        self.labelMolecula = tk.Label(self,text='Molécula')
        self.labelMolecula.place(x=50,y=75)
        
        #Lista desplegable para que el usuario eliga la molécula
        moleculas = set(self.hojaDistribuidor['Molecula'])        
        self.opcionesMoleculas = [mol for mol in moleculas]
        self.opcionesMoleculas.sort()
        self.menuMoleculas = tk.OptionMenu(self, self.moleculaDist, *self.opcionesMoleculas)
        self.menuMoleculas.place(x=50,y=95)
        
        #Label de instrucción del segundo paso
        self.labelMensajePublico = tk.Label(self,text='Paso 2: Elija una presentación')
        self.labelMensajePublico.place(x=50,y=135)
        
        #Variable para el parametro de presentacion
        self.presentacionMolecula = tk.StringVar(self,'Elija una presentación')   
        self.presentacionMolecula.trace_add('write', self.mostrarIdSistema)
        #Texto para la presentacion
        self.labelPresentacion = tk.Label(self,text='Presentación')
        self.labelPresentacion.place(x=50,y=160)
        
        #Lista desplegable para que el usuario eliga la presentacion de la molecula
        self.menuPresentaciones = tk.OptionMenu(self, self.presentacionMolecula, '')
        self.menuPresentaciones.place(x=50,y=180)
        
        #Label de instrucción del tercer paso
        self.labelMensajePublico = tk.Label(self,text='Paso 3: Elija un Id de Sistema')
        self.labelMensajePublico.place(x=300,y=50)
            
        #Variable para el parametro de id sistema
        self.idSistema = tk.StringVar(self,'Elija una id Sistema')   
        self.idSistema.trace_add('write', self.habilitarInformacionAdicional)
        #Texto para la presentacion
        self.labelIdSistema = tk.Label(self,text='ID Sistema')
        self.labelIdSistema.place(x=300,y=75)
        
        #Lista desplegable para que el usuario eliga la presentacion de la molecula
        self.menuIdSistema = tk.OptionMenu(self, self.idSistema, '')
        self.menuIdSistema.place(x=300,y=95)
        
        #Variable para el parametro empresa
        self.companiaDist = tk.StringVar(self,'')
        #Texto para la empresa 
        self.labelCompania = tk.Label(self,text='Empresa')
        self.labelCompania.place(x=300,y=150)
        #Entrada de texto para el procedimiento
        self.entradaCompania = tk.Entry(self, width =30, textvariable=self.companiaDist)
        self.entradaCompania.place(x=285,y=180)
        
        #Label de instrucción del tercer paso
        self.labelMensajePublico = tk.Label(self,text='En caso de no encontrar el Id Sistema llene la siguiente información')
        self.labelMensajePublico.place(x=50,y=230)
        
        #Variable para el parametro empresa
        self.empresaDist = tk.StringVar(self,'')
        #Texto para la empresa 
        self.labelEmpresa = tk.Label(self,text='Empresa')
        self.labelEmpresa.place(x=50,y=260)
        #Entrada de texto para el procedimiento
        self.entradaEmpresa = tk.Entry(self, width =40, textvariable=self.empresaDist,
                                          state=tk.DISABLED)
        self.entradaEmpresa.place(x=110,y=260)
        
        #Texto para la marcapropia 
        self.labelMarca = tk.Label(self,text='Marca Propia')
        self.labelMarca.place(x=50,y=295)

        #Variable para el parametro de cantidad
        self.cantidad=tk.StringVar(self,'0')
        #Texto para la cantidad
        self.labelCantidad = tk.Label(self,text='Cantidad')
        self.labelCantidad.place(x=50,y=330)
        #Entrada de texto para la cantidad
        self.entradaCantidad = tk.Entry(self, width =40, textvariable=self.cantidad,
                                          state=tk.NORMAL)
        self.entradaCantidad.place(x=110,y=330)
        
        #Variable para el parametro de precio
        self.precio=tk.StringVar(self,'0.0')
        #Texto para la cantidad
        self.labelPrecio = tk.Label(self,text='Precio')
        self.labelPrecio.place(x=50,y=370)
        #Entrada de texto para la cantidad
        self.entradaPrecio = tk.Entry(self, width =40, textvariable=self.precio,
                                          state=tk.NORMAL)
        self.entradaPrecio.place(x=110,y=370)
        
        #Texto para el mes
        self.labelMes = tk.Label(self,text='Mes/Meses')
        self.labelMes.place(x=20,y=400)
        
        #Cuadro de texto para mostrar los meses del año
        self.scroll2 = tk.Scrollbar(self, orient=tk.VERTICAL)
        self.listaMesesPublico = tk.Listbox(self, width=30, height=4,selectmode=tk.MULTIPLE,
                                            yscrollcommand=self.scroll2.set,exportselection=False)
        self.scroll2.configure(command=self.listaMesesPublico.yview) 
        self.listaMesesPublico.place(x=175, y=400)
        self.scroll2.place(x=190, y=400)
        
        # Cuadro de texto para mostrar el resumen de los datos ingresados
        self.texto = tk.Text(self, height=5, width=40, wrap='word')
        self.texto.insert(tk.END, "")
        self.texto.config(state='disabled')  # Configurar el widget como solo lectura
        self.texto.place(x=90,y=500)
                
        #Seccion de los RADIOBUTTONS para escoger opciones en la ventana
        
        #Radiobutton para que el usuario distribuidor o retail eliga si ingresa empresa o marca propia
        #Variable para la opción elegida de tipo de medicamento por el usuario
        self.datoTipo = tk.StringVar(self,'ninguno')
        #Eleccion Licitacion
        self.eleccionEmpresa = tk.Radiobutton(self,text="",variable=self.datoTipo,
                                             value="empresa",command=self.habilitacionDeEntradas,
                                             state=tk.DISABLED)
        self.eleccionEmpresa.place(x=370,y=255)
        #Eleccion clave
        self.eleccionMarcaPropia = tk.Radiobutton(self,text="",variable=self.datoTipo,
                                             value="MarcaPropia",command=self.habilitacionDeEntradas,
                                             state=tk.DISABLED)
        self.eleccionMarcaPropia.place(x=370,y=290)
        
        #Radiobutton para que el usuario publico eliga si quiere varios meses o todo el año
        #Variable para la opción elegida de meses por el usuario
        self.datoMes = tk.StringVar(self,'ninguno')
        #Eleccion Licitacion
        self.eleccionMesUnico = tk.Radiobutton(self,text="Por meses",variable=self.datoMes,
                                             value="unicoMes",command=self.habilitacionMeses)
        self.eleccionMesUnico.place(x=375,y=400)
        #Eleccion clave
        self.eleccionCompleto = tk.Radiobutton(self,text="Año \ncompleto",variable=self.datoMes,
                                             value="añoCompleto",command=self.habilitacionMeses)
        self.eleccionCompleto.place(x=375,y=440)
        
        
        #Seccion de los CHECKBUTTONS para elegir los meses
        #Mes de ENERO
        # self.mesEnero = tk.StringVar(self,'')
        # self.checkEnero = tk.Checkbutton(self,text='ENERO',variable=self.mesEnero,onvalue='ENERO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkEnero.place(x=100,y=400)
        
        # #Mes de FEBRERO
        # self.mesFebrero = tk.StringVar(self,'')
        # self.checkFebrero = tk.Checkbutton(self,text='FEBRERO',variable=self.mesFebrero,onvalue='FEBRERO',offvalue='',
        #                                    state=tk.DISABLED)
        # self.checkFebrero.place(x=180,y=400)
        
        # #Mes de MARZO
        # self.mesMarzo = tk.StringVar(self,'')
        # self.checkMarzo = tk.Checkbutton(self,text='MARZO',variable=self.mesMarzo,onvalue='MARZO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkMarzo.place(x=270,y=400)
        
        # #Mes de ABRIL
        # self.mesAbril = tk.StringVar(self,'')
        # self.checkAbril = tk.Checkbutton(self,text='ABRIL',variable=self.mesAbril,onvalue='ABRIL',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkAbril.place(x=100,y=420)
        
        # #Mes de MAYO
        # self.mesMayo = tk.StringVar(self,'')
        # self.checkMayo = tk.Checkbutton(self,text='MAYO',variable=self.mesMayo,onvalue='MAYO',offvalue='',
        #                                 state=tk.DISABLED)
        # self.checkMayo.place(x=180,y=420)
        
        # #Mes de JUNIO
        # self.mesJunio = tk.StringVar(self,'')
        # self.checkJunio = tk.Checkbutton(self,text='JUNIO',variable=self.mesJunio,onvalue='JUNIO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkJunio.place(x=270,y=420)
        
        # #Mes de JULIO
        # self.mesJulio = tk.StringVar(self,'')
        # self.checkJulio = tk.Checkbutton(self,text='JULIO',variable=self.mesJulio,onvalue='JULIO',offvalue='',
        #                                  state=tk.DISABLED)
        # self.checkJulio.place(x=100,y=440)
        
        # #Mes de AGOSTO
        # self.mesAgosto = tk.StringVar(self,'')
        # self.checkAgosto = tk.Checkbutton(self,text='AGOSTO',variable=self.mesAgosto,onvalue='AGOSTO',offvalue='',
        #                                   state=tk.DISABLED)
        # self.checkAgosto.place(x=180,y=440)
        
        # #Mes de SEPTIEMBRE
        # self.mesSeptiembre = tk.StringVar(self,'')
        # self.checkSeptiembre = tk.Checkbutton(self,text='SEPTIEMBRE',variable=self.mesSeptiembre,onvalue='SEPTIEMBRE',offvalue='',
        #                                       state=tk.DISABLED)
        # self.checkSeptiembre.place(x=270,y=440)
        
        # #Mes de OCTUBRE
        # self.mesOctubre = tk.StringVar(self,'')
        # self.checkOctubre = tk.Checkbutton(self,text='OCTUBRE',variable=self.mesOctubre,onvalue='OCTUBRE',offvalue='',
        #                                    state=tk.DISABLED)
        # self.checkOctubre.place(x=100,y=460)
        
        # #Mes de NOVIEMBRE
        # self.mesNoviembre = tk.StringVar(self,'')
        # self.checkNoviembre = tk.Checkbutton(self,text='NOVIEMBRE',variable=self.mesNoviembre,onvalue='NOVIEMBRE',offvalue='',
        #                                      state=tk.DISABLED)
        # self.checkNoviembre.place(x=180,y=460)
        
        # #Mes de DICIEMBRE
        # self.mesDiciembre = tk.StringVar(self,'')
        # self.checkDiciembre = tk.Checkbutton(self,text='DICIEMBRE',variable=self.mesDiciembre,onvalue='DICIEMBRE',offvalue='',
        #                                      state=tk.DISABLED)
        # self.checkDiciembre.place(x=270,y=460)
     
        
        #Seccion de los BUTTONS para realizar acciones dentro de la ventana
        
        #Boton para regresar a la página principal
        self.regresoPublicoInicio = tk.Button(self, text="Regresar al Inicio",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.volverVentanaPrincipal)
        self.regresoPublicoInicio.place(x=330, y=610)
        
        #Boton para mostrar el resumen de datos ingresados
        self.mostrarResumen = tk.Button(self, text="Mostrar resumen de ingreso",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.visualizarResumenEntrada)
        self.mostrarResumen.place(x=40, y=610)
        
        #Boton para guardar la información en el excel de registros
        self.guardarRegistro = tk.Button(self, text="Guardar información",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.guardarInformacion)
        self.guardarRegistro.place(x=185, y=610)
        
#########################################################################################
#################################ESPACIO PARA LAS FUNCIONES##############################
    
    def mostrarPresentaciones(self,*args):
        if self.moleculaDist.get() != 'Elija una molécula':
            filasMoleculas = self.hojaDistribuidor[self.hojaDistribuidor['Molecula']==self.moleculaDist.get()]
            presentaciones = set(filasMoleculas['SKUUnica'])
            opPresentaciones = [p for p in presentaciones]
            print(opPresentaciones)
            menu = self.menuPresentaciones['menu']
            menu.delete(0,'end')
            for presentacion in opPresentaciones:
                menu.add_command(label=presentacion,command=tk._setit(self.presentacionMolecula,presentacion))
                
        else:
            mensaje = 'Elija una molecula primero'
            menu.add_command(label=mensaje,command=tk._setit(self.presentacionMolecula,mensaje))
            
            
    def mostrarIdSistema(self,*args):
        if self.presentacionMolecula.get() != 'Elija una presentación':
            filasId = self.hojaDistribuidor[(self.hojaDistribuidor['Molecula']==self.moleculaDist.get()) & 
                                                   (self.hojaDistribuidor['SKUUnica']==self.presentacionMolecula.get())]
            ids = set(filasId['IDItemSistema'])
            opId = [p for p in ids]
            opId.append('No se encuentra el Id')
            print(opId)
            menu = self.menuIdSistema['menu']
            menu.delete(0,'end')
            for sist in opId:
                menu.add_command(label=sist,command=tk._setit(self.idSistema,sist))
        else:
            mensaje = 'Elija una presentación primero'
            menu.add_command(label=mensaje,command=tk._setit(self.idSistema,mensaje))
            
            
    def habilitarInformacionAdicional(self,*args):
        if self.idSistema.get() == 'No se encuentra el Id':
            self.eleccionEmpresa.config(state=tk.NORMAL)
            self.eleccionMarcaPropia.config(state=tk.NORMAL)
        else:
            self.eleccionEmpresa.config(state=tk.DISABLED)
            self.eleccionMarcaPropia.config(state=tk.DISABLED)
            
            
    def habilitacionDeEntradas(self):
        
        if self.datoTipo.get() == 'empresa':
            self.entradaEmpresa.config(state=tk.NORMAL)
        else:
            self.entradaEmpresa.config(state=tk.DISABLED)
            
    def habilitacionMeses(self):
        
        meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
        
        if self.datoMes.get() == 'unicoMes':
            self.listaMesesPublico.config(state=tk.NORMAL)
            self.listaMesesPublico.delete(0, tk.END)
                     
            # Agregar las claves a la lista
            for mes in meses:
                self.listaMesesPublico.insert(tk.END, mes)
                
        else:
            self.listaMesesPublico.config(state=tk.DISABLED)
            
        # if self.datoMes.get() == 'unicoMes':
        #     self.checkEnero.config(state=tk.NORMAL)
        #     self.checkFebrero.config(state=tk.NORMAL)
        #     self.checkMarzo.config(state=tk.NORMAL)
        #     self.checkAbril.config(state=tk.NORMAL)
        #     self.checkMayo.config(state=tk.NORMAL)
        #     self.checkJunio.config(state=tk.NORMAL)
        #     self.checkJulio.config(state=tk.NORMAL)
        #     self.checkAgosto.config(state=tk.NORMAL)
        #     self.checkSeptiembre.config(state=tk.NORMAL)
        #     self.checkOctubre.config(state=tk.NORMAL)
        #     self.checkNoviembre.config(state=tk.NORMAL)
        #     self.checkDiciembre.config(state=tk.NORMAL)
        # else:
        #     self.checkEnero.config(state=tk.DISABLED)
        #     self.checkFebrero.config(state=tk.DISABLED)
        #     self.checkMarzo.config(state=tk.DISABLED)
        #     self.checkAbril.config(state=tk.DISABLED)
        #     self.checkMayo.config(state=tk.DISABLED)
        #     self.checkJunio.config(state=tk.DISABLED)
        #     self.checkJulio.config(state=tk.DISABLED)
        #     self.checkAgosto.config(state=tk.DISABLED)
        #     self.checkSeptiembre.config(state=tk.DISABLED)
        #     self.checkOctubre.config(state=tk.DISABLED)
        #     self.checkNoviembre.config(state=tk.DISABLED)
        #     self.checkDiciembre.config(state=tk.DISABLED)
            
    def verificarIntegridadDatos(self):
        
        if self.moleculaDist.get()=='Elija una molécula':
            tk.messagebox.showwarning(title=None, message='Elija una molécula')
            return
        elif self.presentacionMolecula.get()=='Elija una presentación':
            tk.messagebox.showwarning(title=None, message='Elija una presentación')
            return
        elif self.idSistema.get()=='Elija una id Sistema':
            tk.messagebox.showwarning(title=None, message='Elija una id Sistema')
            return
        
        if self.idSistema.get() == 'No se encuentra el Id':
            if self.empresaDist.get()=='' and self.datoTipo.get()=='empresa':
                tk.messagebox.showwarning(title=None, message='Ingrese una empresa')
                return
            
        if self.companiaDist.get() == '':
            tk.messagebox.showwarning(title=None, message='Ingrese una empresa')
                  
        if self.cantidad.get()!='0' or self.cantidad.get()=='0':
            try:
                self.cantidadEntero = int(self.cantidad.get())
            except ValueError:
                tk.messagebox.showwarning(title=None, message='El tipo de dato en Cantidad es erróneo')
                return
        else:       
            tk.messagebox.showwarning(title=None, message='Llene toda la información ')
            return
        
        if self.precio.get()!='0.0':
            try:
                self.PrecioFlotante = float(self.precio.get())
            except ValueError:
                tk.messagebox.showwarning(title=None, message='El tipo de precio en Precio es erróneo')
                return
        elif self.precio.get()=='':       
            tk.messagebox.showwarning(title=None, message='Llene toda la información ')
            return
       
        if self.datoMes.get()=='unicoMes':
            # Obtener los meses seleccionados
            seleccionMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
            #Verificar que si se ha seleccionado una clave
            if not seleccionMes:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                return
            
        # if self.datoMes.get()=='unicoMes':
        #     #verificar si hay meses seleeccionados
        #     checks = [self.mesEnero.get(),self.mesFebrero.get(),self.mesMarzo.get(), self.mesAbril.get(),
        #               self.mesMayo.get(),self.mesJunio.get(),self.mesJulio.get(),self.mesAgosto.get(),
        #               self.mesSeptiembre.get(),self.mesOctubre.get(),self.mesNoviembre.get(),
        #               self.mesDiciembre.get()]
            
        #     todosChecks = all(val=='' for val in checks)
            
        #     if todosChecks:
        #         tk.messagebox.showwarning(title=None, message='Seleccione al menos un mes')
        #         return
        
        return True
    
    def obtencionDatosResumen(self):
        
        if self.verificarIntegridadDatos():
            
            #Obtener el dato de molécula
            self.datoMolecula = self.moleculaDist.get()
            
            #Obtener el dato de presentación
            self.datoPresentacion = self.presentacionMolecula.get()
            
            #Obtener la empresa
            self.datoEmpresa = self.companiaDist.get()
            
            #Obtener el id Sistema
            if self.idSistema.get() == 'No se encuentra el Id':
                filasId = self.hojaDistribuidor[(self.hojaDistribuidor['Molecula']==self.moleculaDist.get()) & 
                                                       (self.hojaDistribuidor['SKUUnica']==self.presentacionMolecula.get())]
                ids = set(filasId['IDItemSistema'])
                opId = [p for p in ids]
                opId = opId[0]
                print(opId)
                if self.datoTipo.get()=='empresa':
                    self.datoIdSistema = opId+self.empresaDist.get()
                else:
                    self.datoIdSistema = opId+self.datoTipo.get()
            else: 
                self.datoIdSistema = self.idSistema.get()
            
            
            #Obtener el precio distribuidor
            self.datoPrecio = self.precio.get()
            
            #Obtener la cantidad
            self.datoCantidad = self.cantidad.get()
            
            if self.verificarParametroTiempo():
                return
           
            #Obtener mes o meses 
            if self.datoMes.get()=='unicoMes':
                #Obtener la clave seleccionada 
                self.varMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
                if not self.varMes:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                    return
            else:
                self.varMes = self.datoMes.get()
            # if self.datoMes.get()=='unicoMes':
            #     #Obtener los meses seleccionados
            #     checks = [self.mesEnero.get(),self.mesFebrero.get(),self.mesMarzo.get(), self.mesAbril.get(),
            #               self.mesMayo.get(),self.mesJunio.get(),self.mesJulio.get(),self.mesAgosto.get(),
            #               self.mesSeptiembre.get(),self.mesOctubre.get(),self.mesNoviembre.get(),
            #               self.mesDiciembre.get()]
                
            #     self.varMes = [val for val in checks if val != '']
                
            # else:
            #   self.varMes = self.datoMes.get()
                
                       
            return self.datoMolecula,self.datoPresentacion,self.datoIdSistema,self.datoPrecio,self.datoCantidad,self.varMes, self.datoEmpresa
        
    def verificarParametroTiempo(self):
        if self.datoMes.get() == 'ninguno':
            tk.messagebox.showwarning(title=None, message='Elija un tipo de tiempo')
            return True
        return False
            
    def visualizarResumenEntrada(self):
        mol,pre,sist,prec,cant,me,em=self.obtencionDatosResumen()
        
        self.texto.config(state=tk.NORMAL)
        self.texto.delete(1.0, tk.END)
        self.texto.insert(tk.END, f'Id Sistema: {sist} \n')
        self.texto.insert(tk.END, f'Empresa: {em} \n')
        self.texto.insert(tk.END, f'Molécula: {mol} \n')
        self.texto.insert(tk.END, f'Presentación: {pre} \n')
        self.texto.insert(tk.END, f'Cantidad: {cant} \n')
        self.texto.insert(tk.END, f'Precio: ${prec} \n') 
        if self.datoMes.get()=='unicoMes':
            textoMes = ', '.join(me)
        else:
            textoMes = me
        self.texto.insert(tk.END, f'Mes/Meses: {textoMes} \n')
        self.texto.config(state=tk.DISABLED)
        
    def crearTablaParaGuardar(self):
        mol,pre,sist,prec,cant,me,em=self.obtencionDatosResumen()
        fecha = datetime.now()
        year = fecha.year
        if self.tipoCliente =='distribuidor':
            sector = 'Distribuidor'
        else:
            sector = 'Retail'
            
        if me =='añoCompleto':
            meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                    'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
            dicc = { 'Mes':meses, 'Año':year,'Id Sistema':sist,'Empresa':em,'Molécula':mol, 'SKU':pre,
                    'Cantidad':cant,'Precio':prec,'Sector': sector}
        else:
            dicc = { 'Mes':me, 'Año':year,'Id Sistema':sist,'Empresa':em,'Molécula':mol, 'SKU':pre,
                    'Cantidad':cant,'Precio':prec,'Sector': sector}
        
        dataFrame = pd.DataFrame(dicc)
        print(dataFrame)
        return dataFrame
    
    def guardarInformacionExcel(self):
        dataFrame = self.crearTablaParaGuardar()
        
        try:
            ruta = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
            archivo = os.path.isfile(ruta)
            if archivo:
                with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    libro = pd.ExcelFile(ruta, engine='openpyxl')
                    if 'TablaGeneral' in libro.sheet_names:
                        dfExistente = pd.read_excel(ruta, sheet_name='TablaGeneral')
                        dfCombinado = pd.concat([dfExistente, dataFrame], ignore_index=True)
                        
                        # Eliminar duplicados
                        dfDuplicado = dfCombinado.drop_duplicates(subset=['Mes','Id Sistema'], keep='last')
                        
                        #Filtrar por Sector
                        sector=self.tipoCliente.capitalize()
                        dfSectorFiltrado = dfCombinado[dfCombinado['Sector']==sector]
                        
                        # Verificar duplicados
                        duplicados = dfSectorFiltrado[dfSectorFiltrado.duplicated(subset=['Mes','Id Sistema'], keep=False)]
                        if not duplicados.empty:
                            mensajesDuplicados = duplicados[['Mes','Id Sistema']].drop_duplicates().to_string(index=False)
                            message = f'Los siguientes registros ya están en el excel:\n{mensajesDuplicados}\n y no se ha guardado la información'
                            tk.messagebox.showwarning(title=None, message=message)
                        else:
                            dfCombinado.to_excel(writer, sheet_name='TablaGeneral', index=False)
                            tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
                    else:
                        dataFrame.to_excel(writer, sheet_name='TablaGeneral', index=False)
                        writer.sheets['TablaGeneral'].header_row = 1
                        workbook = writer.book
                        worksheet = writer.sheets['TablaGeneral']
                        
                        # Aplicar formato a los encabezados
                        for cell in worksheet[1]:
                            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
            else:
                dataFrame.to_excel(ruta, sheet_name='TablaGeneral', index=False)
                wb = load_workbook(ruta)
                ws = wb.active
                
                # Definir estilos para los encabezados
                header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                # Aplicar formato a los encabezados
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(ruta)
                tk.messagebox.showinfo(title=None, message='Tabla Creada e Información guardada con éxito')
        except ValueError:
            tk.messagebox.showwarning(title=None, message='Hubo un error al guardar la información')
        
        # try:
        #     ruta = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
        #     archivo = os.path.isfile(ruta)
        #     if archivo:
        #         with pd.ExcelWriter(ruta,engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
        #             libro = pd.ExcelFile(ruta,engine='openpyxl')
        #             if 'TablaGeneral' in libro.sheet_names:
        #                 dfExistente = pd.read_excel(ruta,sheet_name='TablaGeneral')                        
        #                 dfCombinado = pd.concat([dfExistente,dataFrame], ignore_index=True)
        #                 print(dfCombinado)
        #                 sector=self.tipoCliente.capitalize()
        #                 dfSectorFiltrado = dfCombinado[dfCombinado['Sector']==sector]
        #                 dfValoresDuplicados = dfSectorFiltrado[dfSectorFiltrado.duplicated(subset=['Mes','Id Sistema'],keep=False)]
        #                 duplicadosDatos = [dfValoresDuplicados['Mes'].unique(),dfValoresDuplicados['Id Sistema'].unique()]
        #                 mensajesDuplicados = []
        #                 for i in range(1):
        #                     for j in range(len(duplicadosDatos[i])):
        #                         mensajesDuplicados.append(duplicadosDatos[i][j]+" "+str(duplicadosDatos[i+1][0]))
        #                 dfDuplicado = dfSectorFiltrado.drop_duplicates(subset=['Mes','Id Sistema'], keep='first')
        #                 print(dfDuplicado)
        #                 if len(dfSectorFiltrado) == len(dfDuplicado):
        #                     dfDuplicado.to_excel(writer,sheet_name='TablaGeneral',index=False)
        #                     tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
        #                 else:
        #                     message = 'Los siguientes registros ya estan en el excel: \n'+'\n'.join(mensajesDuplicados)+'\n'+'y no se ha guardado la información'
        #                     tk.messagebox.showwarning(title=None, message=message)
        #             else:
        #                 dataFrame.to_excel(writer,sheet_name='TablaGeneral',index=False)
        #                 tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
        #     else:
        #         dataFrame.to_excel(ruta,sheet_name='TablaGeneral',index=False)
        #         path = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
        #         wb = load_workbook(path)
        #         ws = wb.active
        #         # Definir estilos para los encabezados
        #         header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')  # Azul
        #         header_font = Font(color='FFFFFF', bold=True)  # Blanco y negrita
        #         # Aplicar formato a los encabezados
        #         for cell in ws[1]:
        #             cell.fill = header_fill
        #             cell.font = header_font
        #         wb.save(path)
        #         tk.messagebox.showinfo(title=None, message='Tabla Creada e Información guardada con éxito')
        # except ValueError:
        #     tk.messagebox.showwarning(title=None, message='Hubo un error al guardar la información')
        #     return

    def guardarInformacion(self):
        self.verificarIntegridadDatos()
        self.guardarInformacionExcel()
        
    def volverVentanaPrincipal(self):
        self.destroy()
        self.Ventana_Principal.deiconify()

########################################################################################
#######################VENTANA DE MODIFICACION DE INFORMACION###########################    

class VentanaModificacion(tk.Toplevel):
    def __init__(self,VentanaPrincipal):
        super().__init__()
        self.title("Ventana Modificacion")
        self.geometry("600x400")
        self.Ventana_Principal = VentanaPrincipal
        
######################################################################################
################################ESPACIOS PARA LOS ELEMENTOS###########################
        
        #Seccion para los LABEL de la ventana
        
        #Obtener el tipo de cliente 
        self.tipoCliente = self.Ventana_Principal.tipoCliente.get()
        
        #Label del tipo de cliente
        self.labelTitulo = tk.Label(self, text=f"Tipo de cliente: {self.tipoCliente}")
        self.labelTitulo.place(x=225,y=30)
        
        #Label de las instrucciones
        self.labelTitulo = tk.Label(self, text="Eliga la acción que desea realizar con la base de datos")
        self.labelTitulo.place(x=160,y=50)
        
        #Seccion de los BUTTONS para elegir la acción que quiere el usuario
        
        #Boton para ingreso nuevo
        self.botonIngreso = tk.Button(self,text="Actualizar Datos",width=30,
                                      height=10, command=self.abrirVentanaDeActualizacion)
        self.botonIngreso.place(x=50,y=110)
        
        #Boton para modificacion de información
        self.botonModificacion = tk.Button(self,text="Ingresar datos de procedimientos ya existentes",width=30,
                                      height=10,command=self.abrirVentanaDeIngreso,
                                      wraplength=100, justify='center', anchor='center')
        self.botonModificacion.place(x=320,y=110)
        
        #Boton para regresar al Inicio
        self.botonRegresoInicio = tk.Button(self,text="Regresar al Inicio",width=17,height=3,
                                            wraplength=100, justify='center', anchor='center',
                                            command=self.volverVentanaPrincipal)
        self.botonRegresoInicio.place(x=420,y=320)
        
#########################################################################################
#################################ESPACIO PARA LAS FUNCIONES##############################

    #Función para abrir la ventana de modificacion de informacion
    def abrirVentanaDeActualizacion(self):      
        print(self.tipoCliente)
        self.ventana2 = VentanaDeActualizacion(self)
        self.withdraw()
        
    
    def abrirVentanaDeIngreso(self):      
        print(self.tipoCliente)       
        self.ventana2 = VentanaDeIngresoModificacion(self)
        self.withdraw()

    
    def volverVentanaPrincipal(self):
        self.destroy()
        self.Ventana_Principal.deiconify()

#########################################################################################
#######################VENTANA DE ACTUALIZACION DE INFORMACION###########################

class VentanaDeActualizacion(tk.Toplevel):
    def __init__(self,VentanaModificacion):
        super().__init__()
        self.title("Ventana Actualización de datos")
        self.geometry("1100x600")
        self.Ventana_Modificacion = VentanaModificacion
        
        self.tipoCliente = self.Ventana_Modificacion.tipoCliente
        
######################################################################################
################################ESPACIOS PARA LOS ELEMENTOS###########################
        
        #Label del tipo de cliente
        self.labelTitulo = tk.Label(self, text=f"Actualización de datos para el tipo de cliente: {self.tipoCliente}")
        self.labelTitulo.place(x=150,y=30)
        
        if self.tipoCliente == 'publico':
            
            #Seccion para los ENTRY y LABEL  de la sección publico
            
            #Variable para el parametro de procedimiento
            self.procedimientoPublico = tk.StringVar(self,'')     
            #Texto para el procedimiento
            self.labelProc = tk.Label(self,text='Procedimiento/Licitacion')
            self.labelProc.place(x=20,y=130)
            #Entrada de texto para el procedimiento
            self.entradaProc = tk.Entry(self, width =40, textvariable=self.procedimientoPublico,
                                              state=tk.DISABLED)
            self.entradaProc.place(x=175,y=130)
            
            #Variable para el parametro de Clave
            self.clavePublico = tk.StringVar(self,'')     
            #Texto para la clave
            self.labelClave = tk.Label(self,text='Clave')
            self.labelClave.place(x=20,y=170)
            #Entrada de texto para la clave
            self.entradaClave = tk.Entry(self, width =40, textvariable=self.clavePublico,
                                              state=tk.DISABLED)
            self.entradaClave.place(x=175,y=170)
            
            
            #Cuadro de texto para mostrar las claves o procedimientos encontrados
            self.scroll1 = tk.Scrollbar(self, orient=tk.VERTICAL)
            self.listaDatosPublico = tk.Listbox(self, width=30, height=4, yscrollcommand=self.scroll1.set,
                                                selectmode=tk.SINGLE,exportselection=False)
            self.scroll1.configure(command=self.listaDatosPublico.yview)                 
            self.listaDatosPublico.place(x=175, y=210)
            self.scroll1.place(x=190, y=210)
            
            #Variable para el parametro de cantidad
            self.cantidad=tk.StringVar(self,'0')
            #Texto para la cantidad
            self.labelCantidad = tk.Label(self,text='Cantidad')
            self.labelCantidad.place(x=20,y=290)
            #Entrada de texto para la cantidad
            self.entradaCantidad = tk.Entry(self, width =40, textvariable=self.cantidad,
                                              state=tk.DISABLED)
            self.entradaCantidad.place(x=175,y=290)
            
            # Crear un marco para contener el widget Text y las barras de desplazamiento
            self.frame = tk.Frame(self, width=600, height=300)
            self.frame.place(x=550, y=100)
    
            # Crear un widget Text
            self.texto = tk.Text(self.frame, height=20, width=60, wrap='none', font=('Courier New', 10))
            self.texto.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
            # Crear barras de desplazamiento
            scrollbar_y = tk.Scrollbar(self.frame, orient=tk.VERTICAL, command=self.texto.yview)
            scrollbar_x = tk.Scrollbar(self.frame, orient=tk.HORIZONTAL, command=self.texto.xview)
    
            # Configurar las barras de desplazamiento
            self.texto.config(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
    
            # Colocar las barras de desplazamiento
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
            scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

            
            # Cuadro de texto para mostrar el resumen de los datos ingresados
            # self.texto = tk.Text(self, height=30, width=60, wrap='word')
            # self.texto.insert(tk.END, "")
            # self.texto.config(state='disabled')  # Configurar el widget como solo lectura
            # self.texto.place(x=550,y=100)
    
            #Texto para el mes
            self.labelMes = tk.Label(self,text='Paso 2: Eliga un mes/Meses para modificar')
            self.labelMes.place(x=20,y=330)
            
            ########################################################################################################
            #Cuadro de texto para mostrar los meses del año
            self.listaMesesPublico = tk.Listbox(self, width=30, height=4,selectmode=tk.MULTIPLE,
                                               exportselection=False)
            self.listaMesesPublico.place(x=20, y=350)
            meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                    'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
            
            self.listaMesesPublico.delete(0, tk.END)
                         
            # Agregar las claves a la lista
            for mes in meses:
                self.listaMesesPublico.insert(tk.END, mes)
                
            self.listaMesesPublico.config(state=tk.DISABLED)
            
            #Seccion de los RADIOBUTTONS para elegir la clave o procedimiento
            
            #Radiobutton para que el usuario publico eliga si licitacion o clave para ingresar
            #Variable para la opción elegida de busqueda por el usuario
            self.datoIngreso = tk.StringVar(self,'ninguno')
            #Eleccion Licitacion
            self.eleccionProc = tk.Radiobutton(self,text="",variable=self.datoIngreso,
                                                 value="procedimiento",command=self.habilitacionDeEntradas)
            self.eleccionProc.place(x=450,y=130)
            #Eleccion clave
            self.eleccionClave = tk.Radiobutton(self,text="",variable=self.datoIngreso,
                                                 value="clave",command=self.habilitacionDeEntradas)
            self.eleccionClave.place(x=450,y=170)
            
            #Seccion de los BUTTONS para realizar acciones dentro de la ventana
            
            #Boton para mostrar las claves o procedimientos encontrados
            self.mostrarClaves = tk.Button(self, text="Mostrar Claves/Procedimientos",
                                           width=15,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.mostrarClaProc)
            self.mostrarClaves.place(x=370, y=210)
            
            #Boton para regresar a la página principal
            self.regresoPublicoInicio = tk.Button(self, text="Regresar al Inicio",
                                           width=17,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.volverVentanaPrincipal)
            self.regresoPublicoInicio.place(x=330, y=510)
            
            #Boton para mostrar el resumen de registros buscados
            self.mostrarResumen = tk.Button(self, text="Buscar",
                                           width=17,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.mostrarTablaResultados)
            self.mostrarResumen.place(x=40, y=510)
            
            #Boton para guardar la información en el excel de registros
            self.guardarRegistro = tk.Button(self, text="Modificar Registro",
                                           width=17,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.modificarInformacion)
            self.guardarRegistro.place(x=185, y=510)
            
        else:
            
            #Seccion de los ENTRY, LISTBOX y OPTIONMENU para ingresar la información del cliente
            
            #Label de instrucción del primer paso
            self.labelMensajePublico = tk.Label(self,text='Paso 1: Elija una molécula')
            self.labelMensajePublico.place(x=50,y=50)
            
            #Variable para el parametro de molecula
            self.moleculaDist = tk.StringVar(self,'Elija una molécula') 
            self.moleculaDist.trace_add('write', self.mostrarPresentaciones)
            #Texto para la molecula
            self.labelMolecula = tk.Label(self,text='Molécula')
            self.labelMolecula.place(x=50,y=75)
            
            #Lista desplegable para que el usuario eliga la molécula
            dfs = self.cargarBaseDeDatos()
            filas = dfs[dfs['Sector']==self.tipoCliente.capitalize()]
            moleculas = set(filas['Molécula'])  
            print(moleculas)
            self.opcionesMoleculas = [mol for mol in moleculas]
            self.opcionesMoleculas.sort()
            self.menuMoleculas = tk.OptionMenu(self, self.moleculaDist, *self.opcionesMoleculas)
            self.menuMoleculas.place(x=50,y=95)
            
            #Label de instrucción del segundo paso
            self.labelMensajePublico = tk.Label(self,text='Paso 2: Elija una presentación')
            self.labelMensajePublico.place(x=50,y=135)
            
            #Variable para el parametro de presentacion
            self.presentacionMolecula = tk.StringVar(self,'Elija una presentación')   
            self.presentacionMolecula.trace_add('write', self.mostrarIdSistema)
            #Texto para la presentacion
            self.labelPresentacion = tk.Label(self,text='Presentación')
            self.labelPresentacion.place(x=50,y=160)
            
            #Lista desplegable para que el usuario eliga la presentacion de la molecula
            self.menuPresentaciones = tk.OptionMenu(self, self.presentacionMolecula, '')
            self.menuPresentaciones.place(x=50,y=180)
            
            #Label de instrucción del tercer paso
            self.labelMensajePublico = tk.Label(self,text='Paso 3: Elija un Id de Sistema')
            self.labelMensajePublico.place(x=300,y=50)
                
            #Variable para el parametro de id sistema
            self.idSistema = tk.StringVar(self,'Elija una id Sistema')   
            #Texto para la presentacion
            self.labelIdSistema = tk.Label(self,text='ID Sistema')
            self.labelIdSistema.place(x=300,y=75)
            
            #Lista desplegable para que el usuario eliga la presentacion de la molecula
            self.menuIdSistema = tk.OptionMenu(self, self.idSistema, '')
            self.menuIdSistema.place(x=300,y=95)
            
            #Variable para el parametro de cantidad
            self.cantidad=tk.StringVar(self,'0')
            #Texto para la cantidad
            self.labelCantidad = tk.Label(self,text='Cantidad')
            self.labelCantidad.place(x=50,y=240)
            #Entrada de texto para la cantidad
            self.entradaCantidad = tk.Entry(self, width =40, textvariable=self.cantidad,
                                              state=tk.DISABLED)
            self.entradaCantidad.place(x=115,y=240)
            
            #Variable para el parametro de precio
            self.precio=tk.StringVar(self,'0.0')
            #Texto para la cantidad
            self.labelPrecio = tk.Label(self,text='Precio')
            self.labelPrecio.place(x=50,y=280)
            #Entrada de texto para la cantidad
            self.entradaPrecio = tk.Entry(self, width =40, textvariable=self.precio,
                                              state=tk.DISABLED)
            self.entradaPrecio.place(x=115,y=280)
            
            # Crear un marco para contener el widget Text y las barras de desplazamiento
            self.frame = tk.Frame(self, width=600, height=300)
            self.frame.place(x=550, y=100)
    
            # Crear un widget Text
            self.texto = tk.Text(self.frame, height=20, width=60, wrap='none', font=('Courier New', 10))
            self.texto.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
            # Crear barras de desplazamiento
            scrollbar_y = tk.Scrollbar(self.frame, orient=tk.VERTICAL, command=self.texto.yview)
            scrollbar_x = tk.Scrollbar(self.frame, orient=tk.HORIZONTAL, command=self.texto.xview)
    
            # Configurar las barras de desplazamiento
            self.texto.config(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
    
            # Colocar las barras de desplazamiento
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
            scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
            
            #Texto para el mes
            self.labelMes = tk.Label(self,text='Paso 2: Eliga un mes/Meses para modificar')
            self.labelMes.place(x=50,y=330)
            
            ########################################################################################################
            #Cuadro de texto para mostrar los meses del año
            self.listaMesesPublico = tk.Listbox(self, width=30, height=4,selectmode=tk.MULTIPLE,
                                               exportselection=False)
            self.listaMesesPublico.place(x=50, y=350)
            meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                    'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
            
            self.listaMesesPublico.delete(0, tk.END)
                         
            # Agregar las claves a la lista
            for mes in meses:
                self.listaMesesPublico.insert(tk.END, mes)
                
            self.listaMesesPublico.config(state=tk.DISABLED)
            
            #Seccion de los BUTTONS para realizar acciones dentro de la ventana
            
            #Boton para regresar a la página principal
            self.regresoPublicoInicio = tk.Button(self, text="Regresar al Inicio",
                                           width=17,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.volverVentanaPrincipal)
            self.regresoPublicoInicio.place(x=330, y=510)
            
            #Boton para mostrar el resumen de registros buscados
            self.mostrarResumen = tk.Button(self, text="Buscar",
                                           width=17,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.mostrarTablaResultados)
            self.mostrarResumen.place(x=40, y=510)
            
            #Boton para modificar la información en el excel de registros
            self.guardarRegistro = tk.Button(self, text="Modificar Registro",
                                           width=17,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.modificarInformacion)
            self.guardarRegistro.place(x=185, y=510)
            
            
#########################################################################################
#################################ESPACIO PARA LAS FUNCIONES##############################
    
    def cargarBaseDeDatos(self):
        try:
            url = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
            dfs = pd.read_excel(url)
            print('base cargada')
        except Exception as e:
            print(f'Error al cargar el archivo de {url} verifique la direccion: {str(e)}')
            
        return dfs
    
    def mostrarClaProc(self):           
            
        self.parametroBusqueda = self.capturaDeParametroBusqueda()
        print(self.parametroBusqueda)
        
        if self.datoIngreso.get() == 'procedimiento':
            if not self.parametroBusqueda:
                tk.messagebox.showwarning(title=None, message='Por favor ingrese una licitación.')
                return
            clavesEncontradas = self.obtenerClavesProcedimientos(self.parametroBusqueda)
            self.listaDatosPublico.delete(0, tk.END)
            
            # Agregar las claves a la lista
            for clave in clavesEncontradas:
                self.listaDatosPublico.insert(tk.END, clave)
        else:
            if not self.parametroBusqueda:
                tk.messagebox.showwarning(title=None, message='Por favor ingrese una clave')
                return
        
            procedimientosEncontrados = self.obtenerClavesProcedimientos(self.parametroBusqueda)
            
            self.listaDatosPublico.delete(0, tk.END)
            
            # Agregar las claves a la lista
            for procedimiento in procedimientosEncontrados:
                self.listaDatosPublico.insert(tk.END, procedimiento)
        
    def obtenerClavesProcedimientos(self, pBusqueda):
        
        self.dfs = self.cargarBaseDeDatos()
        
        if self.datoIngreso.get() == 'procedimiento':
            self.filasProc = self.dfs[self.dfs['Procedimiento/Licitación'] == pBusqueda]
            clavesProc = self.filasProc['Clave'].tolist()
            clavesProc = set(clavesProc)
            return clavesProc
        else:
            self.filasProc = self.dfs[self.dfs['Clave'] == pBusqueda]
            ProcClav = self.filasProc['Procedimiento/Licitación'].tolist()
            ProcClav = set(ProcClav)
            return ProcClav
        
    def habilitacionDeEntradas(self):
        if self.datoIngreso.get() == 'procedimiento':
            self.entradaClave.config(state=tk.DISABLED)
            self.entradaProc.config(state=tk.NORMAL)
        else:
            self.entradaClave.config(state=tk.NORMAL)
            self.entradaProc.config(state=tk.DISABLED)
        
    def capturaDeParametroBusqueda(self):
        if self.verificarParametroBusqueda():
            return
        
        if self.datoIngreso.get() == 'procedimiento':
            self.proc_str = self.procedimientoPublico.get()
            return self.proc_str 
        else:
            self.clave_str = self.clavePublico.get()
            return self.clave_str
        
    def verificarParametroBusqueda(self):
        if self.datoIngreso.get() == '':
            tk.messagebox.showwarning(title=None, message='Elija un parametro de busqueda')
            return True
        return False
    
    def verificarParametroTiempo(self):
        if self.datoMes.get() == 'ninguno':
            tk.messagebox.showwarning(title=None, message='Elija un tipo de tiempo')
            return True
        return False
    
    def mostrarTablaResultados(self):
        if self.verificarIntegridadDatosBusqueda():
            
            if self.tipoCliente == 'publico':
        
                # Obtener la clave o procedimiento seleccionada
                seleccion = self.listaDatosPublico.curselection()
                #Verificar que si se ha seleccionado una clave
                if not seleccion:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                    return 
                
                #Obtener el procedimiento y clave
                if self.datoIngreso.get()=='procedimiento':
                    self.datoProcedimiento = self.procedimientoPublico.get()
                    #Obtener la clave seleccionada 
                    self.datoClave = self.listaDatosPublico.get(seleccion[0])
        
                else:
                    #Obtener la clave seleccionada 
                    self.datoProcedimiento = self.listaDatosPublico.get(seleccion[0])
                    self.datoClave = self.clavePublico.get()
                
                #Obtener los registros de la busqueda
                self.filas = self.dfs[(self.dfs['Procedimiento/Licitación'] == self.datoProcedimiento) & (self.dfs['Clave'] == self.datoClave)]
                self.filas.drop(columns=self.filas.columns[7:], inplace=True)
                self.filasPublico = self.filas
                # Convertir el DataFrame a una cadena formateada
                text_data = self.filasPublico.to_string(index=False, header=True, formatters={'Mes': '{:<10}'.format,
                                                                      'Año': '{:<10}'.format,
                                                                      'Clave': '{:<10}'.format,
                                                                      'Licitación': '{:<10}'.format,
                                                                      'Cantidad': '{:<10}'.format,
                                                                      'Precio': '{:<10}'.format,
                                                                      'Sector': '{:<10}'.format})
                self.texto.config(state=tk.NORMAL)
                self.texto.delete(1.0, tk.END)
                self.texto.insert(tk.END, text_data)       
                self.texto.config(state=tk.DISABLED)
                
                self.listaMesesPublico.config(state=tk.NORMAL)
                self.entradaCantidad.config(state=tk.NORMAL)
            else:
                #Obtener el dato de molécula
                self.datoMolecula = self.moleculaDist.get()
                
                #Obtener el dato de presentación
                self.datoPresentacion = self.presentacionMolecula.get()
                
                #Obtener el id Sistema               
                self.datoIdSistema = self.idSistema.get()
                              
                
                #Obtener los registros de la busqueda
                self.filas = self.dfs2[(self.dfs2['Molécula'] == self.datoMolecula) & (self.dfs2['SKU'] == self.datoPresentacion) & (self.dfs2['Id Sistema'] == self.datoIdSistema)]
                self.filas.drop(columns=self.filas.columns[2:4], inplace=True)
                self.filasDist = self.filas
                # Convertir el DataFrame a una cadena formateada
                text_data = self.filasDist.to_string(index=False, header=True, formatters={'Mes': '{:<10}'.format,
                                                                      'Año': '{:<10}'.format,
                                                                      'Cantidad': '{:<10}'.format,
                                                                      'Precio': '{:<10}'.format,
                                                                      'Sector': '{:<10}'.format,
                                                                      'Id Sistema': '{:<10}'.format,
                                                                      'Empresa': '{:<10}'.format,
                                                                      'Molécula': '{:<10}'.format,
                                                                      'SKU': '{:<10}'.format})
                self.texto.config(state=tk.NORMAL)
                self.texto.delete(1.0, tk.END)
                self.texto.insert(tk.END, text_data)       
                self.texto.config(state=tk.DISABLED)
                
                self.listaMesesPublico.config(state=tk.NORMAL)
                self.entradaCantidad.config(state=tk.NORMAL)
                self.entradaPrecio.config(state=tk.NORMAL)
            
    def mostrarPresentaciones(self,*args):
        
        self.dfs2 = self.cargarBaseDeDatos()
        if self.moleculaDist.get() != 'Elija una molécula':
            filasMoleculas = self.dfs2[self.dfs2['Molécula']==self.moleculaDist.get()]
            presentaciones = set(filasMoleculas['SKU'])
            opPresentaciones = [p for p in presentaciones]
            print(opPresentaciones)
            menu = self.menuPresentaciones['menu']
            menu.delete(0,'end')
            for presentacion in opPresentaciones:
                menu.add_command(label=presentacion,command=tk._setit(self.presentacionMolecula,presentacion))
                
        else:
            mensaje = 'Elija una molecula primero'
            menu.add_command(label=mensaje,command=tk._setit(self.presentacionMolecula,mensaje))
            
    def mostrarIdSistema(self,*args):
        if self.presentacionMolecula.get() != 'Elija una presentación':
            filasId = self.dfs2[(self.dfs2['Molécula']==self.moleculaDist.get()) & 
                                                   (self.dfs2['SKU']==self.presentacionMolecula.get())]
            ids = set(filasId['Id Sistema'])
            opId = [p for p in ids]
            print(opId)
            menu = self.menuIdSistema['menu']
            menu.delete(0,'end')
            for sist in opId:
                menu.add_command(label=sist,command=tk._setit(self.idSistema,sist))
        else:
            mensaje = 'Elija una presentación primero'
            menu.add_command(label=mensaje,command=tk._setit(self.idSistema,mensaje))
            
    
    
    def verificarIntegridadDatosBusqueda(self):
        
        if self.tipoCliente == 'publico':
        
            if self.datoIngreso.get()=='procedimiento' and self.procedimientoPublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese un procedimiento o licitación ')
                return
            elif self.datoIngreso.get()=='clave' and self.clavePublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese una clave')
                return
            
            
            # Obtener la clave o procedimiento seleccionada
            seleccion = self.listaDatosPublico.curselection()
            #Verificar que si se ha seleccionado una clave
            if not seleccion:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                return
                      

        else:
            if self.moleculaDist.get()=='Elija una molécula':
                tk.messagebox.showwarning(title=None, message='Elija una molécula')
                return
            elif self.presentacionMolecula.get()=='Elija una presentación':
                tk.messagebox.showwarning(title=None, message='Elija una presentación')
                return
            elif self.idSistema.get()=='Elija una id Sistema':
                tk.messagebox.showwarning(title=None, message='Elija una id Sistema')
                return
    
        
        return True
    
    def verificarIntegridadDatos(self):
        
        if self.tipoCliente == 'publico':
        
            if self.datoIngreso.get()=='procedimiento' and self.procedimientoPublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese un procedimiento o licitación ')
                return
            elif self.datoIngreso.get()=='clave' and self.clavePublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese una clave')
                return
            
            
            # Obtener la clave o procedimiento seleccionada
            seleccion = self.listaDatosPublico.curselection()
            #Verificar que si se ha seleccionado una clave
            if not seleccion:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                return
                      
            if self.cantidad.get()!='0' or self.cantidad.get()=='0':
                try:
                    self.cantidadEntero = int(self.cantidad.get())
                except ValueError:
                    tk.messagebox.showwarning(title=None, message='El tipo de dato en Cantidad es erróneo')
                    return
            else:       
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
            
    
            # Obtener los meses seleccionados
            seleccionMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
            #Verificar que si se ha seleccionado una clave
            if not seleccionMes:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                return
        else:
            if self.moleculaDist.get()=='Elija una molécula':
                tk.messagebox.showwarning(title=None, message='Elija una molécula')
                return
            elif self.presentacionMolecula.get()=='Elija una presentación':
                tk.messagebox.showwarning(title=None, message='Elija una presentación')
                return
            elif self.idSistema.get()=='Elija una id Sistema':
                tk.messagebox.showwarning(title=None, message='Elija una id Sistema')
                return
                      
            if self.cantidad.get()!='0' or self.cantidad.get()=='0':
                try:
                    self.cantidadEntero = int(self.cantidad.get())
                except ValueError:
                    tk.messagebox.showwarning(title=None, message='El tipo de dato en Cantidad es erróneo')
                    return
            else:       
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
            
            if self.precio.get()!='0.0':
                try:
                    self.PrecioFlotante = float(self.precio.get())
                except ValueError:
                    tk.messagebox.showwarning(title=None, message='El tipo de precio en Precio es erróneo')
                    return
            elif self.precio.get()=='':       
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
           
            # Obtener los meses seleccionados
            seleccionMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
            #Verificar que si se ha seleccionado una clave
            if not seleccionMes:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                return
        
        return True
    
    def modificarInformacion(self):
        if self.verificarIntegridadDatos():
            
            self.varMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
            mesesRegistrados = self.filas['Mes'].tolist()
            for mes in self.varMes:
                if mes not in mesesRegistrados:
                    tk.messagebox.showwarning(title=None, message=f'El mes {mes} no se encuentra registrado en la base')
                    return

                
            if self.tipoCliente == 'publico':
                cliente = 'público'
                for mes in self.varMes:
                    filtro = (self.dfs['Procedimiento/Licitación'] == self.datoProcedimiento) & \
                             (self.dfs['Clave'] == self.datoClave) & (self.dfs['Mes'] == mes) & \
                             (self.dfs['Sector'] == cliente.capitalize())
                            
                    self.dfs.loc[filtro,'Cantidad'] = self.cantidad.get()
                    
            else:
                for mes in self.varMes:
                    filtro = (self.dfs2['Molécula'] == self.datoMolecula) & \
                             (self.dfs2['SKU'] == self.datoPresentacion) & \
                             (self.dfs2['Id Sistema'] == self.datoIdSistema) & (self.dfs2['Mes'] == mes) & \
                             (self.dfs2['Sector'] == self.tipoCliente.capitalize())  
                                
                    self.dfs2.loc[filtro,'Cantidad'] = self.cantidad.get()                   
                    if self.precio.get() != '0.0' or self.precio.get() == '':
                        self.dfs2.loc[filtro,'Precio'] = self.precio.get()
    
            self.guardarCambiosModificacion()
            
            
    def guardarCambiosModificacion(self):
        url ='C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
        with pd.ExcelWriter(url,engine='openpyxl', mode='a', if_sheet_exists = 'replace') as writer:
            if self.tipoCliente == 'publico':
                self.dfs.to_excel(writer,sheet_name='TablaGeneral', index=False)
                tk.messagebox.showinfo(title=None, message='Información modificada con éxito')
            else:
                self.dfs2.to_excel(writer,sheet_name='TablaGeneral', index=False)
                tk.messagebox.showinfo(title=None, message='Información modificada con éxito')
    
    def volverVentanaPrincipal(self):
        self.destroy()
        self.Ventana_Modificacion.deiconify()
        
        
#########################################################################################
#######################VENTANA DE ACTUALIZACION DE INFORMACION###########################

class VentanaDeIngresoModificacion(tk.Toplevel):
    def __init__(self,VentanaModificacion):
        super().__init__()
        self.title("Ventana Ingreso de datos")
        self.geometry("1100x650")
        self.Ventana_Modificacion = VentanaModificacion
        
        self.tipoCliente = self.Ventana_Modificacion.tipoCliente
        
        #Label para la ventana de tipo cliente publico
        self.labelMensajePublico = tk.Label(self,text=f'Ingreso de datos del tipo de cliente {self.tipoCliente}')
        self.labelMensajePublico.place(x=130,y=10)
        
        if self.tipoCliente == 'publico':
            #Seccion de los ENTRY y LISTBOX para ingresar la información del cliente
            
            #Variable para el parametro de procedimiento
            self.procedimientoPublico = tk.StringVar(self,'')     
            #Texto para el procedimiento
            self.labelProc = tk.Label(self,text='Procedimiento/Licitacion')
            self.labelProc.place(x=20,y=130)
            #Entrada de texto para el procedimiento
            self.entradaProc = tk.Entry(self, width =40, textvariable=self.procedimientoPublico,
                                              state=tk.DISABLED)
            self.entradaProc.place(x=175,y=130)
            
            #Variable para el parametro de Clave
            self.clavePublico = tk.StringVar(self,'')     
            #Texto para la clave
            self.labelClave = tk.Label(self,text='Clave')
            self.labelClave.place(x=20,y=170)
            #Entrada de texto para la clave
            self.entradaClave = tk.Entry(self, width =40, textvariable=self.clavePublico,
                                              state=tk.DISABLED)
            self.entradaClave.place(x=175,y=170)
            
            
            #Cuadro de texto para mostrar las claves o procedimientos encontrados
            self.scroll1 = tk.Scrollbar(self, orient=tk.VERTICAL)
            self.listaDatosPublico = tk.Listbox(self, width=30, height=4, yscrollcommand=self.scroll1.set,
                                                selectmode=tk.SINGLE,exportselection=False)
            self.scroll1.configure(command=self.listaDatosPublico.yview)                 
            self.listaDatosPublico.place(x=175, y=210)
            self.scroll1.place(x=190, y=210)
            
            #Variable para el parametro de cantidad
            self.cantidad=tk.StringVar(self,'0')
            #Texto para la cantidad
            self.labelCantidad = tk.Label(self,text='Cantidad')
            self.labelCantidad.place(x=20,y=290)
            #Entrada de texto para la cantidad
            self.entradaCantidad = tk.Entry(self, width =40, textvariable=self.cantidad,
                                              state=tk.NORMAL)
            self.entradaCantidad.place(x=175,y=290)
            
            #Texto para el mes
            self.labelMes = tk.Label(self,text='Mes/Meses')
            self.labelMes.place(x=20,y=330)
            
            ########################################################################################################
            #Cuadro de texto para mostrar los meses del año
            self.scroll2 = tk.Scrollbar(self, orient=tk.VERTICAL)
            self.listaMesesPublico = tk.Listbox(self, width=30, height=4,selectmode=tk.MULTIPLE,
                                                yscrollcommand=self.scroll2.set,exportselection=False)
            self.scroll2.configure(command=self.listaMesesPublico.yview) 
            self.listaMesesPublico.place(x=175, y=330)
            self.scroll2.place(x=190, y=330)
            ########################################################################################################
            
            
            # Cuadro de texto para mostrar el resumen de los datos ingresados
            self.texto = tk.Text(self, height=5, width=40, wrap='word')
            self.texto.insert(tk.END, "")
            self.texto.config(state='disabled')  # Configurar el widget como solo lectura
            self.texto.place(x=90,y=440)
            
            #Seccion de los RADIOBUTTONS para elegir la clave o procedimiento
            
            #Radiobutton para que el usuario publico eliga si licitacion o clave para ingresar
            #Variable para la opción elegida de busqueda por el usuario
            self.datoIngreso = tk.StringVar(self,'ninguno')
            #Eleccion Licitacion
            self.eleccionProc = tk.Radiobutton(self,text="",variable=self.datoIngreso,
                                                 value="procedimiento",command=self.habilitacionDeEntradas)
            self.eleccionProc.place(x=450,y=130)
            #Eleccion clave
            self.eleccionClave = tk.Radiobutton(self,text="",variable=self.datoIngreso,
                                                 value="clave",command=self.habilitacionDeEntradas)
            self.eleccionClave.place(x=450,y=170)
            
            #Radiobutton para que el usuario publico eliga si quiere varios meses o todo el año
            #Variable para la opción elegida de meses por el usuario
            self.datoMes = tk.StringVar(self,'ninguno')
            #Eleccion Licitacion
            self.eleccionMesUnico = tk.Radiobutton(self,text="Por meses",variable=self.datoMes,
                                                 value="unicoMes",command=self.habilitacionMeses)
            self.eleccionMesUnico.place(x=375,y=330)
            #Eleccion clave
            self.eleccionCompleto = tk.Radiobutton(self,text="Año \ncompleto",variable=self.datoMes,
                                                 value="añoCompleto",command=self.habilitacionMeses)
            self.eleccionCompleto.place(x=375,y=360)
            
            #Seccion de los BUTTONS para realizar acciones dentro de la ventana
            
            #Boton para mostrar las claves o procedimientos encontrados
            self.mostrarClaves = tk.Button(self, text="Mostrar Claves/Procedimientos",
                                           width=15,height=3,wraplength=100, justify='center', 
                                           anchor='center',command=self.mostrarClaProc)
            self.mostrarClaves.place(x=370, y=210)
            
        else:
            
            if self.tipoCliente =='distribuidor':
                mensaje = 'Ingreso de datos del tipo de cliente DISTRIBUIDOR'
            else:
                mensaje = 'Ingreso de datos del tipo de cliente RETAIL'
                
            self.dfs=self.cargarBaseDeDatos()

            #Label para la ventana de tipo cliente distribuidor o Retail
            self.labelMensajePublico = tk.Label(self,text=mensaje)
            self.labelMensajePublico.place(x=130,y=10)
            
            #Seccion de los ENTRY, LISTBOX y OPTIONMENU para ingresar la información del cliente
            
            #Label de instrucción del primer paso
            self.labelMensajePublico = tk.Label(self,text='Paso 1: Elija una molécula')
            self.labelMensajePublico.place(x=50,y=50)
            
            #Variable para el parametro de molecula
            self.moleculaDist = tk.StringVar(self,'Elija una molécula') 
            self.moleculaDist.trace_add('write', self.mostrarPresentaciones)
            #Texto para la molecula
            self.labelMolecula = tk.Label(self,text='Molécula')
            self.labelMolecula.place(x=50,y=75)
            
            #Lista desplegable para que el usuario eliga la molécula
            dfs = self.cargarBaseDeDatos()
            filas = dfs[dfs['Sector']==self.tipoCliente.capitalize()]
            moleculas = set(filas['Molécula'])  
            print(moleculas)       
            self.opcionesMoleculas = [mol for mol in moleculas]
            self.opcionesMoleculas.sort()
            self.menuMoleculas = tk.OptionMenu(self, self.moleculaDist, *self.opcionesMoleculas)
            self.menuMoleculas.place(x=50,y=95)
            
            #Label de instrucción del segundo paso
            self.labelMensajePublico = tk.Label(self,text='Paso 2: Elija una presentación')
            self.labelMensajePublico.place(x=50,y=135)
            
            #Variable para el parametro de presentacion
            self.presentacionMolecula = tk.StringVar(self,'Elija una presentación')   
            self.presentacionMolecula.trace_add('write', self.mostrarIdSistema)
            #Texto para la presentacion
            self.labelPresentacion = tk.Label(self,text='Presentación')
            self.labelPresentacion.place(x=50,y=160)
            
            #Lista desplegable para que el usuario eliga la presentacion de la molecula
            self.menuPresentaciones = tk.OptionMenu(self, self.presentacionMolecula, '')
            self.menuPresentaciones.place(x=50,y=180)
            
            #Label de instrucción del tercer paso
            self.labelMensajePublico = tk.Label(self,text='Paso 3: Elija un Id de Sistema')
            self.labelMensajePublico.place(x=300,y=50)
                
            #Variable para el parametro de id sistema
            self.idSistema = tk.StringVar(self,'Elija una id Sistema')   
            #Texto para la presentacion
            self.labelIdSistema = tk.Label(self,text='ID Sistema')
            self.labelIdSistema.place(x=300,y=75)
            
            #Lista desplegable para que el usuario eliga la presentacion de la molecula
            self.menuIdSistema = tk.OptionMenu(self, self.idSistema, '')
            self.menuIdSistema.place(x=300,y=95)
            
            #Variable para el parametro empresa
            self.companiaDist = tk.StringVar(self,'')
            #Texto para la empresa 
            self.labelCompania = tk.Label(self,text='Empresa')
            self.labelCompania.place(x=300,y=150)
            #Entrada de texto para el procedimiento
            self.entradaCompania = tk.Entry(self, width =30, textvariable=self.companiaDist)
            self.entradaCompania.place(x=285,y=180)

            #Variable para el parametro de cantidad
            self.cantidad=tk.StringVar(self,'0')
            #Texto para la cantidad
            self.labelCantidad = tk.Label(self,text='Cantidad')
            self.labelCantidad.place(x=50,y=250)
            #Entrada de texto para la cantidad
            self.entradaCantidad = tk.Entry(self, width =40, textvariable=self.cantidad,
                                              state=tk.NORMAL)
            self.entradaCantidad.place(x=110,y=250)
            
            #Variable para el parametro de precio
            self.precio=tk.StringVar(self,'0.0')
            #Texto para la cantidad
            self.labelPrecio = tk.Label(self,text='Precio')
            self.labelPrecio.place(x=50,y=300)
            #Entrada de texto para la cantidad
            self.entradaPrecio = tk.Entry(self, width =40, textvariable=self.precio,
                                              state=tk.NORMAL)
            self.entradaPrecio.place(x=110,y=300)
            
            #Texto para el mes
            self.labelMes = tk.Label(self,text='Mes/Meses')
            self.labelMes.place(x=20,y=350)
            
            #Cuadro de texto para mostrar los meses del año
            self.scroll2 = tk.Scrollbar(self, orient=tk.VERTICAL)
            self.listaMesesPublico = tk.Listbox(self, width=30, height=4,selectmode=tk.MULTIPLE,
                                                yscrollcommand=self.scroll2.set,exportselection=False)
            self.scroll2.configure(command=self.listaMesesPublico.yview) 
            self.listaMesesPublico.place(x=175, y=350)
            self.scroll2.place(x=190, y=350)
            
            # Cuadro de texto para mostrar el resumen de los datos ingresados
            self.texto = tk.Text(self, height=5, width=40, wrap='word')
            self.texto.insert(tk.END, "")
            self.texto.config(state='disabled')  # Configurar el widget como solo lectura
            self.texto.place(x=90,y=440)
                    
            #Seccion de los RADIOBUTTONS para escoger opciones en la ventana
            
            #Radiobutton para que el usuario publico eliga si quiere varios meses o todo el año
            #Variable para la opción elegida de meses por el usuario
            self.datoMes = tk.StringVar(self,'ninguno')
            #Eleccion Licitacion
            self.eleccionMesUnico = tk.Radiobutton(self,text="Por meses",variable=self.datoMes,
                                                 value="unicoMes",command=self.habilitacionMeses)
            self.eleccionMesUnico.place(x=375,y=350)
            #Eleccion clave
            self.eleccionCompleto = tk.Radiobutton(self,text="Año \ncompleto",variable=self.datoMes,
                                                 value="añoCompleto",command=self.habilitacionMeses)
            self.eleccionCompleto.place(x=375,y=380)
            
        
        # Crear un marco para contener el widget Text y las barras de desplazamiento
        self.frame = tk.Frame(self, width=600, height=300)
        self.frame.place(x=550, y=100)

        # Crear un widget Text
        self.texto2 = tk.Text(self.frame, height=20, width=60, wrap='none', font=('Courier New', 10))
        self.texto2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Crear barras de desplazamiento
        scrollbar_y = tk.Scrollbar(self.frame, orient=tk.VERTICAL, command=self.texto2.yview)
        scrollbar_x = tk.Scrollbar(self.frame, orient=tk.HORIZONTAL, command=self.texto2.xview)

        # Configurar las barras de desplazamiento
        self.texto2.config(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        # Colocar las barras de desplazamiento
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
            
        #Seccion de los BUTTONS para realizar acciones dentro de la ventana
        
        #Boton para mostrar el resumen de datos ingresados
        self.mostrarResumen = tk.Button(self, text="Mostrar resumen de ingreso",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.visualizarResumenEntrada)
        self.mostrarResumen.place(x=185, y=550)
        
        #Boton para guardar la información en el excel de registros
        self.guardarRegistro = tk.Button(self, text="Guardar información",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.guardarInformacion)
        self.guardarRegistro.place(x=330, y=550)
        
        #Boton para regresar a la página principal
        self.regresoPublicoInicio = tk.Button(self, text="Regresar al Inicio",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.volverVentanaPrincipal)
        self.regresoPublicoInicio.place(x=475, y=550)
        
        #Boton para mostrar el resumen de registros buscados
        self.mostrarRegistros = tk.Button(self, text="Buscar",
                                       width=17,height=3,wraplength=100, justify='center', 
                                       anchor='center',command=self.mostrarTablaResultados)
        self.mostrarRegistros.place(x=40, y=550)
        
#########################################################################################
#################################ESPACIO PARA LAS FUNCIONES##############################
    def cargarBaseDeDatos(self):
        try:
            url = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
            dfs = pd.read_excel(url)
            print('base cargada')
        except Exception as e:
            print(f'Error al cargar el archivo de {url} verifique la direccion: {str(e)}')
            
        return dfs
    
    
    def mostrarClaProc(self):       
        self.parametroBusqueda = self.capturaDeParametroBusqueda()
        print(self.parametroBusqueda)
        
        if self.datoIngreso.get() == 'procedimiento':
            if not self.parametroBusqueda:
                tk.messagebox.showwarning(title=None, message='Por favor ingrese una licitación.')
                return
            clavesEncontradas = self.obtenerClavesProcedimientos(self.parametroBusqueda)
            self.listaDatosPublico.delete(0, tk.END)
            
            # Agregar las claves a la lista
            for clave in clavesEncontradas:
                self.listaDatosPublico.insert(tk.END, clave)
        else:
            if not self.parametroBusqueda:
                tk.messagebox.showwarning(title=None, message='Por favor ingrese una clave')
                return
        
            procedimientosEncontrados = self.obtenerClavesProcedimientos(self.parametroBusqueda)
            
            self.listaDatosPublico.delete(0, tk.END)
            
            # Agregar las claves a la lista
            for procedimiento in procedimientosEncontrados:
                self.listaDatosPublico.insert(tk.END, procedimiento)
        
    def obtenerClavesProcedimientos(self, pBusqueda):
        
        self.dfs = self.cargarBaseDeDatos()
        
        if self.datoIngreso.get() == 'procedimiento':
            self.filasProc = self.dfs[self.dfs['Procedimiento/Licitación'] == pBusqueda]
            clavesProc = self.filasProc['Clave'].tolist()
            clavesProc = set(clavesProc)
            return clavesProc
        else:
            self.filasProc = self.dfs[self.dfs['Clave'] == pBusqueda]
            ProcClav = self.filasProc['Procedimiento/Licitación'].tolist()
            ProcClav = set(ProcClav)
            return ProcClav
        
    def obtenerPrecio(self, dElegido):
        print(self.parametroBusqueda)
        
        if self.datoIngreso.get() == 'procedimiento':
            self.filasClav = self.dfs[(self.dfs['Clave'] == dElegido) & (self.dfs['Procedimiento/Licitación']==self.parametroBusqueda)]
            precioClav = self.filasClav['Precio'].tolist()
            print(precioClav)
            contador = Counter(precioClav)
            precioClav,frecuencia = contador.most_common(1)[0]
        else:
            self.filasClav = self.dfs[(self.dfs['Clave'] == self.parametroBusqueda) & (self.dfs['Procedimiento/Licitación']==dElegido)]
            precioClav = self.filasClav['Precio'].tolist()
            print(precioClav)
            contador = Counter(precioClav)
            precioClav,frecuencia = contador.most_common(1)[0]

        return precioClav
      
    
    def capturaDeParametroBusqueda(self):
        if self.verificarParametroBusqueda():
            return
        
        if self.datoIngreso.get() == 'procedimiento':
            self.proc_str = self.procedimientoPublico.get()
            return self.proc_str 
        else:
            self.clave_str = self.clavePublico.get()
            return self.clave_str 
        
    def verificarParametroBusqueda(self):
        if self.datoIngreso.get() == '':
            tk.messagebox.showwarning(title=None, message='Elija un parametro de busqueda')
            return True
        return False
    
    def verificarParametroTiempo(self):
        if self.datoMes.get() == 'ninguno':
            tk.messagebox.showwarning(title=None, message='Elija un tipo de tiempo')
            return True
        return False
    
    def verificarIntegridadDatos(self):
        
        if self.tipoCliente == 'publico':
            
            if self.datoIngreso.get()=='ninguno' or self.datoMes.get() == 'ninguno':
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
            elif self.datoIngreso.get()=='procedimiento' and self.procedimientoPublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese un procedimiento o licitación')
                return
            elif self.datoIngreso.get()=='clave' and self.clavePublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese una clave')
                return
            
            # Obtener la clave o procedimiento seleccionada
            seleccion = self.listaDatosPublico.curselection()
            #Verificar que si se ha seleccionado una clave
            if not seleccion:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                return
                      
            if self.cantidad.get()!='0' or self.cantidad.get()=='0':
                try:
                    self.cantidadEntero = int(self.cantidad.get())
                except ValueError:
                    tk.messagebox.showwarning(title=None, message='El tipo de dato en Cantidad es erróneo')
                    return
            else:       
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
            ######################################################################################################
            if self.datoMes.get()=='unicoMes':
                # Obtener los meses seleccionados
                seleccionMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
                #Verificar que si se ha seleccionado una clave
                if not seleccionMes:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                    return
            
            return True
        else:
            if self.moleculaDist.get()=='Elija una molécula':
                tk.messagebox.showwarning(title=None, message='Elija una molécula')
                return
            elif self.presentacionMolecula.get()=='Elija una presentación':
                tk.messagebox.showwarning(title=None, message='Elija una presentación')
                return
            elif self.idSistema.get()=='Elija una id Sistema':
                tk.messagebox.showwarning(title=None, message='Elija una id Sistema')
                return
            
            if self.idSistema.get() == 'No se encuentra el Id':
                if self.empresaDist.get()=='' and self.datoTipo.get()=='empresa':
                    tk.messagebox.showwarning(title=None, message='Ingrese una empresa')
                    return
            
            if self.companiaDist.get() == '':
                tk.messagebox.showwarning(title=None, message='Ingrese una empresa')
            
            if self.cantidad.get()!='0' or self.cantidad.get()=='0':
                try:
                    self.cantidadEntero = int(self.cantidad.get())
                except ValueError:
                    tk.messagebox.showwarning(title=None, message='El tipo de dato en Cantidad es erróneo')
                    return
            else:       
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
            
            if self.precio.get()!='0.0':
                try:
                    self.PrecioFlotante = float(self.precio.get())
                except ValueError:
                    tk.messagebox.showwarning(title=None, message='El tipo de precio en Precio es erróneo')
                    return
            elif self.precio.get()=='':       
                tk.messagebox.showwarning(title=None, message='Llene toda la información ')
                return
           
            if self.datoMes.get()=='unicoMes':
                # Obtener los meses seleccionados
                seleccionMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
                #Verificar que si se ha seleccionado una clave
                if not seleccionMes:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                    return
            return True
    
    
    def obtencionDatosResumen(self):
        
        if self.verificarIntegridadDatos():
            
            if self.tipoCliente == 'publico':
        
                # Obtener la clave o procedimiento seleccionada
                seleccion = self.listaDatosPublico.curselection()
                #Verificar que si se ha seleccionado una clave
                if not seleccion:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                    return 
                
                #Obtener el procedimiento y clave
                if self.datoIngreso.get()=='procedimiento':
                    self.datoProcedimiento = self.procedimientoPublico.get()
                    #Obtener la clave seleccionada 
                    self.datoClave = self.listaDatosPublico.get(seleccion[0])
        
                else:
                    #Obtener la clave seleccionada 
                    self.datoProcedimiento = self.listaDatosPublico.get(seleccion[0])
                    self.datoClave = self.clavePublico.get()
                    
                print(self.datoProcedimiento)
                print(self.datoClave)
                
                #Obtener el precio publico
                if self.datoIngreso.get() == 'procedimiento':
                    self.datoPrecio = self.obtenerPrecio(self.datoClave)
                else:
                    self.datoPrecio = self.obtenerPrecio(self.datoProcedimiento)
                
                print(self.datoPrecio)
                
                #Obtener la cantidad
                self.datoCantidad = self.cantidad.get()
                
                print(self.datoCantidad)
                
                if self.verificarParametroTiempo():
                    return
                #########################################################################################################
                #Obtener mes o meses 
                if self.datoMes.get()=='unicoMes':
                    #Obtener la clave seleccionada 
                    self.varMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
                    if not self.varMes:
                        tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                        return
                else:
                    self.varMes = self.datoMes.get()
                    
                print(self.varMes)
                    
                           
                return self.datoProcedimiento,self.datoClave,self.datoPrecio,self.datoCantidad,self.varMes
            else:
                #Obtener el dato de molécula
                self.datoMolecula = self.moleculaDist.get()
                
                #Obtener el dato de presentación
                self.datoPresentacion = self.presentacionMolecula.get()
                
                #Obtener el id Sistema
                if self.idSistema.get() == 'No se encuentra el Id':
                    filasId = self.hojaDistribuidor[(self.hojaDistribuidor['Molecula']==self.moleculaDist.get()) & 
                                                           (self.hojaDistribuidor['SKUUnica']==self.presentacionMolecula.get())]
                    ids = set(filasId['IDItemSistema'])
                    opId = [p for p in ids]
                    opId = opId[0]
                    print(opId)
                    if self.datoTipo.get()=='empresa':
                        self.datoIdSistema = opId+self.empresaDist.get()
                    else:
                        self.datoIdSistema = opId+self.datoTipo.get()
                else: 
                    self.datoIdSistema = self.idSistema.get()
                    
                self.datoEmpresa = self.companiaDist.get()
                
                
                #Obtener el precio distribuidor
                self.datoPrecio = self.precio.get()
                
                #Obtener la cantidad
                self.datoCantidad = self.cantidad.get()
                
                if self.verificarParametroTiempo():
                    return
               
                #Obtener mes o meses 
                if self.datoMes.get()=='unicoMes':
                    #Obtener la clave seleccionada 
                    self.varMes = [self.listaMesesPublico.get(i) for i in self.listaMesesPublico.curselection()]
                    if not self.varMes:
                        tk.messagebox.showwarning(title=None, message='Por favor seleccione un mes o meses')
                        return
                else:
                    self.varMes = self.datoMes.get()
                           
                return self.datoMolecula,self.datoPresentacion,self.datoIdSistema,self.datoPrecio,self.datoCantidad,self.varMes,self.datoEmpresa
            
    def visualizarResumenEntrada(self):
        
        if self.tipoCliente == 'publico':
            proc,cla,pre,can,me=self.obtencionDatosResumen()
            
            self.texto.config(state=tk.NORMAL)
            self.texto.delete(1.0, tk.END)
            self.texto.insert(tk.END, f'Licitación/Procedimiento: {proc} \n')
            self.texto.insert(tk.END, f'Clave: {cla} \n')        
            self.texto.insert(tk.END, f'Cantidad: {can} \n')
            self.texto.insert(tk.END, f'Precio: ${pre} \n')
            if self.datoMes.get()=='unicoMes':
                textoMes = ', '.join(me)
            else:
                textoMes = me
            self.texto.insert(tk.END, f'Mes/Meses: {textoMes} \n')
            self.texto.config(state=tk.DISABLED)
        else:
            mol,pre,sist,prec,cant,me,em=self.obtencionDatosResumen()
            
            self.texto.config(state=tk.NORMAL)
            self.texto.delete(1.0, tk.END)
            self.texto.insert(tk.END, f'Id Sistema: {sist} \n')
            self.texto.insert(tk.END, f'Empresa: {em} \n')
            self.texto.insert(tk.END, f'Molécula: {mol} \n')
            self.texto.insert(tk.END, f'Presentación: {pre} \n')
            self.texto.insert(tk.END, f'Cantidad: {cant} \n')
            self.texto.insert(tk.END, f'Precio: ${prec} \n') 
            if self.datoMes.get()=='unicoMes':
                textoMes = ', '.join(me)
            else:
                textoMes = me
            self.texto.insert(tk.END, f'Mes/Meses: {textoMes} \n')
            self.texto.config(state=tk.DISABLED)
        
    def mostrarPresentaciones(self,*args):
        if self.moleculaDist.get() != 'Elija una molécula':
            filasMoleculas = self.dfs[self.dfs['Molécula']==self.moleculaDist.get()]
            presentaciones = set(filasMoleculas['SKU'])
            opPresentaciones = [p for p in presentaciones]
            print(opPresentaciones)
            menu = self.menuPresentaciones['menu']
            menu.delete(0,'end')
            for presentacion in opPresentaciones:
                menu.add_command(label=presentacion,command=tk._setit(self.presentacionMolecula,presentacion))
                
        else:
            mensaje = 'Elija una molecula primero'
            menu.add_command(label=mensaje,command=tk._setit(self.presentacionMolecula,mensaje))
            
            
    def mostrarIdSistema(self,*args):
        if self.presentacionMolecula.get() != 'Elija una presentación':
            filasId = self.dfs[(self.dfs['Molécula']==self.moleculaDist.get()) & 
                                                   (self.dfs['SKU']==self.presentacionMolecula.get())]
            ids = set(filasId['Id Sistema'])
            opId = [p for p in ids]
            menu = self.menuIdSistema['menu']
            menu.delete(0,'end')
            for sist in opId:
                menu.add_command(label=sist,command=tk._setit(self.idSistema,sist))
        else:
            mensaje = 'Elija una presentación primero'
            menu.add_command(label=mensaje,command=tk._setit(self.idSistema,mensaje))
            
            
    def habilitacionMeses(self):
        
        meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
        
        if self.datoMes.get() == 'unicoMes':
            self.listaMesesPublico.config(state=tk.NORMAL)
            self.listaMesesPublico.delete(0, tk.END)
                     
            # Agregar las claves a la lista
            for mes in meses:
                self.listaMesesPublico.insert(tk.END, mes)
                
        else:
            self.listaMesesPublico.config(state=tk.DISABLED)
        
    def mostrarTablaResultados(self):
        if self.verificarIntegridadDatosBusqueda():
            
            if self.tipoCliente == 'publico':
        
                # Obtener la clave o procedimiento seleccionada
                seleccion = self.listaDatosPublico.curselection()
                #Verificar que si se ha seleccionado una clave
                if not seleccion:
                    tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                    return 
                
                #Obtener el procedimiento y clave
                if self.datoIngreso.get()=='procedimiento':
                    self.datoProcedimiento = self.procedimientoPublico.get()
                    #Obtener la clave seleccionada 
                    self.datoClave = self.listaDatosPublico.get(seleccion[0])
        
                else:
                    #Obtener la clave seleccionada 
                    self.datoProcedimiento = self.listaDatosPublico.get(seleccion[0])
                    self.datoClave = self.clavePublico.get()
                
                #Obtener los registros de la busqueda
                self.filas = self.dfs[(self.dfs['Procedimiento/Licitación'] == self.datoProcedimiento) & (self.dfs['Clave'] == self.datoClave)]
                self.filas.drop(columns=self.filas.columns[7:], inplace=True)
                self.filasPublico = self.filas
                # Convertir el DataFrame a una cadena formateada
                text_data = self.filasPublico.to_string(index=False, header=True, formatters={'Mes': '{:<10}'.format,
                                                                      'Año': '{:<10}'.format,
                                                                      'Clave': '{:<10}'.format,
                                                                      'Licitación': '{:<10}'.format,
                                                                      'Cantidad': '{:<10}'.format,
                                                                      'Precio': '{:<10}'.format,
                                                                      'Sector': '{:<10}'.format})
                self.texto2.config(state=tk.NORMAL)
                self.texto2.delete(1.0, tk.END)
                self.texto2.insert(tk.END, text_data)       
                self.texto2.config(state=tk.DISABLED)
            else:
                #Obtener el dato de molécula
                self.datoMolecula = self.moleculaDist.get()
                
                #Obtener el dato de presentación
                self.datoPresentacion = self.presentacionMolecula.get()
                
                #Obtener el id Sistema               
                self.datoIdSistema = self.idSistema.get()
                
                #Obtener la empresa
                self.datoCompania = self.companiaDist.get()
                              
                
                #Obtener los registros de la busqueda
                self.filas = self.dfs[(self.dfs['Molécula'] == self.datoMolecula) & (self.dfs['SKU'] == self.datoPresentacion) & (self.dfs['Id Sistema'] == self.datoIdSistema)]
                self.filas.drop(columns=self.filas.columns[2:4], inplace=True)
                self.filasDist = self.filas
                # Convertir el DataFrame a una cadena formateada
                text_data = self.filasDist.to_string(index=False, header=True, formatters={'Mes': '{:<10}'.format,
                                                                      'Año': '{:<10}'.format,
                                                                      'Cantidad': '{:<10}'.format,
                                                                      'Precio': '{:<10}'.format,
                                                                      'Sector': '{:<10}'.format,
                                                                      'Id Sistema': '{:<10}'.format,
                                                                      'Empresa':'{:<10}'.format,
                                                                      'Molécula': '{:<10}'.format,
                                                                      'SKU': '{:<10}'.format})
                self.texto2.config(state=tk.NORMAL)
                self.texto2.delete(1.0, tk.END)
                self.texto2.insert(tk.END, text_data)       
                self.texto2.config(state=tk.DISABLED)
                
                self.listaMesesPublico.config(state=tk.NORMAL)
                self.entradaCantidad.config(state=tk.NORMAL)
                self.entradaPrecio.config(state=tk.NORMAL)
                
                
    def verificarIntegridadDatosBusqueda(self):
        
        if self.tipoCliente == 'publico':
        
            if self.datoIngreso.get()=='procedimiento' and self.procedimientoPublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese un procedimiento o licitación ')
                return
            elif self.datoIngreso.get()=='clave' and self.clavePublico.get()=='':
                tk.messagebox.showwarning(title=None, message='Ingrese una clave')
                return
            
            
            # Obtener la clave o procedimiento seleccionada
            seleccion = self.listaDatosPublico.curselection()
            #Verificar que si se ha seleccionado una clave
            if not seleccion:
                tk.messagebox.showwarning(title=None, message='Por favor seleccione una clave o licitacion.')
                return
        return True
        
    def crearTablaParaGuardar(self):
        
        if self.tipoCliente == 'publico':
            
            proc,cla,pre,can,me=self.obtencionDatosResumen()
            fecha = datetime.now()
            year = fecha.year
            if me =='añoCompleto':
                meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                        'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
                dicc = { 'Mes':meses, 'Año':year,'Clave':cla,'Procedimiento/Licitación':proc,
                        'Cantidad':can,'Precio':pre,'Sector':'Público'}
            else:
                dicc = { 'Mes':me, 'Año':year,'Clave':cla,'Procedimiento/Licitación':proc,
                        'Cantidad':can,'Precio':pre,'Sector':'Público'}
            
            dataFrame = pd.DataFrame(dicc)
            print(dataFrame)
            return dataFrame
        else:
            mol,pre,sist,prec,cant,me=self.obtencionDatosResumen()
            fecha = datetime.now()
            year = fecha.year
            if self.tipoCliente =='distribuidor':
                sector = 'Distribuidor'
            else:
                sector = 'Retail'
                
            if me =='añoCompleto':
                meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                        'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
                dicc = { 'Mes':meses, 'Año':year,'Id Sistema':sist,'Molécula':mol, 'SKU':pre,
                        'Cantidad':cant,'Precio':prec,'Sector': sector}
            else:
                dicc = { 'Mes':me, 'Año':year,'Id Sistema':sist,'Molécula':mol, 'SKU':pre,
                        'Cantidad':cant,'Precio':prec,'Sector': sector}
            
            dataFrame = pd.DataFrame(dicc)
            print(dataFrame)
            return dataFrame
    
    def guardarInformacionExcel(self):
        dataFrame = self.crearTablaParaGuardar()
        
        if self.tipoCliente == 'publico':
            try:
                ruta = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
                archivo = os.path.isfile(ruta)    
                with pd.ExcelWriter(ruta,engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
                    libro = pd.ExcelFile(ruta,engine='openpyxl')
                    if 'TablaGeneral' in libro.sheet_names:
                        dfExistente = pd.read_excel(ruta,sheet_name='TablaGeneral')                        
                        dfCombinado = pd.concat([dfExistente,dataFrame], ignore_index=True)
                        print(dfCombinado)
                        sector='Público'
                        dfSectorFiltrado = dfCombinado[dfCombinado['Sector']==sector]
                        dfValoresDuplicados = dfSectorFiltrado[dfSectorFiltrado.duplicated(subset=['Mes','Clave','Procedimiento/Licitación'],keep=False)]
                        duplicadosDatos = [dfValoresDuplicados['Mes'].unique(),dfValoresDuplicados['Clave'].unique(),dfValoresDuplicados['Procedimiento/Licitación'].unique()]
                        mensajesDuplicados = []
                        for i in range(1):
                            for j in range(len(duplicadosDatos[i])):
                                mensajesDuplicados.append(duplicadosDatos[i][j]+" "+duplicadosDatos[i+1][0]+' '+duplicadosDatos[i+2][0])
                                
                        dfDuplicado = dfSectorFiltrado.drop_duplicates(subset=['Mes','Clave','Procedimiento/Licitación'], keep='first')
                        print(dfDuplicado)
                        if len(dfSectorFiltrado) == len(dfDuplicado):
                            dfDuplicado.to_excel(writer,sheet_name='TablaGeneral',index=False)
                            tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
                        else:
                            message = 'Los siguientes registros ya estan en el excel: \n'+'\n'.join(mensajesDuplicados)+'\n'+'y no se ha guardado la información'
                            tk.messagebox.showwarning(title=None, message=message)
                    else:
                        dataFrame.to_excel(writer,sheet_name='TablaGeneral',index=False)
                        writer.sheets['TablaGeneral'].header_row = 1
                        workbook  = writer.book
                        worksheet = writer.sheets['TablaGeneral']
                        # Aplicar formato a los encabezados
                        for cell in worksheet[1]:
                            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
            except ValueError:
                tk.messagebox.showwarning(title=None, message='Hubo un error al guardar la información')
                return
        else:
            try:
                ruta = 'C:/Users/Mayra Herrera/OneDrive - Laboratorios Vanquish SA de CV/Documentos/2024_Base_General_LE_V01.xlsx'
                archivo = os.path.isfile(ruta)
                if archivo:
                    with pd.ExcelWriter(ruta,engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
                        libro = pd.ExcelFile(ruta,engine='openpyxl')
                        if 'TablaGeneral' in libro.sheet_names:
                            dfExistente = pd.read_excel(ruta,sheet_name='TablaGeneral')                        
                            dfCombinado = pd.concat([dfExistente,dataFrame], ignore_index=True)
                            print(dfCombinado)
                            
                            sector=self.tipoCliente.capitalize()
                            dfSectorFiltrado = dfCombinado[dfCombinado['Sector']==sector]
                            dfValoresDuplicados = dfSectorFiltrado[dfSectorFiltrado.duplicated(subset=['Mes','Id Sistema'],keep=False)]
                            duplicadosDatos = [dfValoresDuplicados['Mes'].unique(),dfValoresDuplicados['Id Sistema'].unique()]
                            mensajesDuplicados = []
                            for i in range(1):
                                for j in range(len(duplicadosDatos[i])):
                                    mensajesDuplicados.append(duplicadosDatos[i][j]+" "+str(duplicadosDatos[i+1][0]))
                            dfDuplicado = dfSectorFiltrado.drop_duplicates(subset=['Mes','Id Sistema'], keep='first')
                            if len(dfSectorFiltrado) == len(dfDuplicado):
                                dfDuplicado.to_excel(writer,sheet_name='TablaGeneral',index=False)
                                tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
                            else:
                                message = 'Los siguientes registros ya estan en el excel: \n'+'\n'.join(mensajesDuplicados)+'\n'+'y no se ha guardado la información'
                                tk.messagebox.showwarning(title=None, message=message)
                        else:
                            dataFrame.to_excel(writer,sheet_name='TablaGeneral',index=False)
                            tk.messagebox.showinfo(title=None, message='Información guardada con éxito')
            except ValueError:
                tk.messagebox.showwarning(title=None, message='Hubo un error al guardar la información')
                return
            
            
    def guardarInformacion(self):
        self.verificarIntegridadDatos()
        self.guardarInformacionExcel()
              
    def volverVentanaPrincipal(self):
        self.destroy()
        self.Ventana_Modificacion.deiconify()
        
    def habilitacionDeEntradas(self):
        if self.datoIngreso.get() == 'procedimiento':
            self.entradaClave.config(state=tk.DISABLED)
            self.entradaProc.config(state=tk.NORMAL)
        else:
            self.entradaClave.config(state=tk.NORMAL)
            self.entradaProc.config(state=tk.DISABLED)
            
    def habilitacionMeses(self):
        ######################################################################################
        meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO',
                'SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
        
        if self.datoMes.get() == 'unicoMes':
            self.listaMesesPublico.config(state=tk.NORMAL)
            self.listaMesesPublico.delete(0, tk.END)
                     
            # Agregar las claves a la lista
            for mes in meses:
                self.listaMesesPublico.insert(tk.END, mes)
                
        else:
            self.listaMesesPublico.config(state=tk.DISABLED)

if __name__ == "__main__":
    app = VentanaPrincipal()
    app.mainloop()